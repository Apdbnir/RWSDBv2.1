"""
Модуль менеджера базы данных для системы управления базами данных

Содержит класс DatabaseManager, который реализует функциональность
для работы с PostgreSQL базами данных.
"""

import sys
import os
import json
from datetime import datetime
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView,
                             QPushButton, QComboBox, QLineEdit, QLabel, QGroupBox,
                             QFormLayout, QScrollArea, QSplitter, QFrame, QFileDialog,
                             QMessageBox, QDialog, QInputDialog, QCheckBox, QSpinBox, QTextEdit)
from PyQt6.QtCore import Qt, QAbstractTableModel, QModelIndex
from PyQt6.QtGui import QAction, QKeySequence
import openpyxl
from openpyxl.styles import Font, PatternFill
import psycopg2  # Библиотека для работы с PostgreSQL
from psycopg2 import sql  # фМодуль для безопасной работы с SQL-запросами в PostgreSQL
from utils.db_logging import db_logger  # Модуль для логирования операций с базой данных


class DatabaseManager:
    """
    Класс менеджера базы данных

    Обеспечивает подключение и выполнение операций с PostgreSQL базой данных.
    """

    def __init__(self, db_type='postgresql', connection_params=None):
        """
        Инициализация менеджера базы данных

        Args:
            db_type (str): Тип базы данных ('postgresql')
            connection_params (dict): Параметры подключения к базе данных
        """
        if db_type != 'postgresql':
            raise ValueError("Only PostgreSQL is supported")

        self.db_type = db_type  # Тип базы данных
        self.connection_params = connection_params or {}  # Параметры подключения
        self.connection = None  # Объект соединения
        self.cursor = None  # Объект курсора для выполнения запросов
        self.in_transaction = False  # Флаг, указывающий, находимся ли мы в транзакции
        self.transaction_stack = []  # Стек для хранения информации о транзакциях
        self.connect()  # Выполняем подключение при создании экземпляра

    def connect(self):
        """
        Установление соединения с PostgreSQL базой данных
        """
        try:
            # Подключаемся к PostgreSQL базе данных
            print(f"Попытка подключения к PostgreSQL базе данных: {self.connection_params.get('database', 'unknown')}")
            self.connection = psycopg2.connect(**self.connection_params)
            self.cursor = self.connection.cursor()
            print("Подключение к PostgreSQL успешно!")

            # Логируем успешную попытку подключения
            db_logger.log_connection_attempt(self.db_type, self.connection_params, success=True)
            return True

        except Exception as e:
            error_msg = str(e)
            print(f"Ошибка подключения: {error_msg}")
            # Логируем неудачную попытку подключения
            db_logger.log_connection_attempt(self.db_type, self.connection_params, success=False, error_msg=error_msg)
            return False

    def execute_query(self, query, params=None):
        """
        Выполнение SQL-запроса

        Args:
            query (str): SQL-запрос для выполнения
            params (tuple): Параметры для подстановки в запрос (для защиты от SQL-инъекций)

        Returns:
            list or None: Результаты запроса (для SELECT) или None (для других запросов)
        """
        try:
            print(f"Выполнение запроса: {query[:100]}...")  # Первые 100 символов запроса
            if params:
                print(f"С параметрами: {params}")
                # Выполняем запрос с параметрами (для защиты от SQL-инъекций)
                self.cursor.execute(query, params)
            else:
                # Выполняем запрос без параметров
                self.cursor.execute(query)

            if query.strip().upper().startswith('SELECT'):
                # Если это SELECT-запрос, возвращаем результаты
                result = self.cursor.fetchall()
                print(f"Запрос вернул {len(result) if result else 0} строк")
                # Логируем успешное выполнение запроса
                db_logger.log_query_execution(query, params, success=True)
                return result
            else:
                # Для других запросов (INSERT, UPDATE, DELETE) применяем изменения
                # Только если не находимся в активной транзакции
                if not self.in_transaction:
                    self.connection.commit()
                print("Запрос, не возвращающий данные, выполнен успешно")
                # Логируем успешное выполнение запроса
                db_logger.log_query_execution(query, params, success=True)
                return None

        except psycopg2.InternalError as e:
            error_msg = str(e)
            print(f"Внутренняя ошибка PostgreSQL: {error_msg}")
            # Логируем неудачное выполнение запроса
            db_logger.log_query_execution(query, params, success=False, error_msg=error_msg)

            # Для внутренних ошибок PostgreSQL, таких как "current transaction is aborted",
            # нужно пересоздать соединение
            self.connect()

            raise e
        except Exception as e:
            error_msg = str(e)
            print(f"Выполнение запроса не удалось: {error_msg}")
            # Логируем неудачное выполнение запроса
            db_logger.log_query_execution(query, params, success=False, error_msg=error_msg)

            # Всегда делаем rollback при ошибке, чтобы очистить состояние транзакции
            try:
                self.connection.rollback()
            except:
                # Если rollback не удается, возможно соединение повреждено
                # Попробуем восстановить соединение
                try:
                    self.connect()
                except:
                    pass

            raise e

    def get_tables(self):
        """
        Получение списка таблиц в базе данных

        Returns:
            list: Список имен таблиц
        """
        print(f"Получение списка таблиц из {self.db_type} базы данных")
        try:
            # Запрос к PostgreSQL для получения имен таблиц
            self.cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_schema='public';")

            # Извлекаем имена таблиц из результата
            tables = [table[0] for table in self.cursor.fetchall()]
            print(f"Найдено {len(tables)} таблиц: {tables}")
            return tables
        except psycopg2.InternalError as e:
            print(f"Внутренняя ошибка PostgreSQL при получении таблиц: {str(e)}")
            # Восстанавливаем соединение при внутренней ошибке
            self.connect()
            return []
        except Exception as e:
            print(f"Ошибка получения таблиц: {str(e)}")
            return []

    def get_table_info(self, table_name):
        """
        Получение информации о структуре таблицы

        Args:
            table_name (str): Имя таблицы

        Returns:
            list: Список словарей с информацией о столбцах
        """
        # SQL-запрос для получения информации о столбцах из PostgreSQL
        query = """
        SELECT
            c.column_name,
            c.data_type,
            c.is_nullable,
            c.column_default,
            CASE WHEN pk.constraint_type = 'PRIMARY KEY' THEN 1 ELSE 0 END AS is_primary_key
        FROM information_schema.columns c
        LEFT JOIN (
            SELECT kcu.column_name, tc.constraint_type
            FROM information_schema.table_constraints tc
            JOIN information_schema.key_column_usage kcu
            ON tc.constraint_name = kcu.constraint_name
            WHERE tc.table_name = %s AND tc.constraint_type = 'PRIMARY KEY'
        ) pk ON c.column_name = pk.column_name
        WHERE c.table_name = %s
        ORDER BY c.ordinal_position;
        """
        try:
            self.cursor.execute(query, (table_name, table_name))
            columns = self.cursor.fetchall()
            return [{'name': col[0], 'type': col[1], 'not_null': col[2] == 'NO', 'default': col[3], 'primary_key': col[4] == 1} for col in columns]
        except psycopg2.InternalError as e:
            print(f"Внутренняя ошибка PostgreSQL при получении информации о таблице {table_name}: {str(e)}")
            # Восстанавливаем соединение при внутренней ошибке
            self.connect()
            raise e

    def table_exists(self, table_name):
        """
        Проверяет, существует ли таблица в базе данных

        Args:
            table_name (str): Имя таблицы для проверки

        Returns:
            bool: True, если таблица существует, иначе False
        """
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                SELECT tablename
                FROM pg_tables
                WHERE schemaname = 'public' AND tablename = %s;
            """, (table_name,))
            result = cursor.fetchone()
            return result is not None
        except psycopg2.InternalError as e:
            print(f"Внутренняя ошибка PostgreSQL при проверке существования таблицы {table_name}: {str(e)}")
            # Восстанавливаем соединение при внутренней ошибке
            self.connect()
            return False

    def close(self):
        """
        Закрытие соединения с базой данных
        """
        if self.connection:
            self.connection.close()

    def begin_transaction(self):
        """
        Начало новой транзакции

        Returns:
            bool: True, если транзакция успешно начата, иначе False
        """
        if self.in_transaction:
            print("Предупреждение: Уже находимся в транзакции. Вложенные транзакции не поддерживаются в этой реализации.")
            return False
        try:
            # Начало транзакции в PostgreSQL
            self.cursor.execute("BEGIN")
            self.in_transaction = True
            return True
        except psycopg2.InternalError as e:
            print(f"Внутренняя ошибка PostgreSQL при начале транзакции: {e}")
            # Восстанавливаем соединение при внутренней ошибке
            self.connect()
            return False
        except Exception as e:
            print(f"Ошибка начала транзакции: {e}")
            return False

    def commit_transaction(self):
        """
        Фиксация текущей транзакции

        Returns:
            bool: True, если транзакция успешно зафиксирована, иначе False
        """
        if not self.in_transaction:
            print("Предупреждение: Нет активной транзакции для фиксации.")
            return False
        try:
            self.connection.commit()
            self.in_transaction = False
            return True
        except psycopg2.InternalError as e:
            print(f"Внутренняя ошибка PostgreSQL при фиксации транзакции: {e}")
            # Восстанавливаем соединение при внутренней ошибке
            self.connect()
            return False
        except Exception as e:
            print(f"Ошибка фиксации транзакции: {e}")
            return False

    def rollback_transaction(self):
        """
        Откат текущей транзакции

        Returns:
            bool: True, если транзакция успешно откачена, иначе False
        """
        if not self.in_transaction:
            print("Предупреждение: Нет активной транзакции для отката.")
            return False
        try:
            self.connection.rollback()
            self.in_transaction = False
            return True
        except psycopg2.InternalError as e:
            print(f"Внутренняя ошибка PostgreSQL при откате транзакции: {e}")
            # Восстанавливаем соединение при внутренней ошибке
            self.connect()
            return False
        except Exception as e:
            print(f"Ошибка отката транзакции: {e}")
            return False

    def execute_transaction(self, queries_list):
        """
        Выполнение списка запросов как одной транзакции

        Args:
            queries_list (list): Список запросов для выполнения в транзакции

        Returns:
            list: Результаты выполнения запросов
        """
        if not isinstance(queries_list, list):
            print("Ошибка: Запросы должны быть предоставлены в виде списка.")
            return False

        # Проверяем, не находимся ли мы уже в транзакции
        if self.in_transaction:
            print("Ошибка: Уже находимся в транзакции. Нельзя начать новую транзакцию внутри активной.")
            return False

        try:
            # Начинаем транзакцию
            if not self.begin_transaction():
                raise Exception("Не удалось начать транзакцию")

            results = []
            for query_info in queries_list:
                if isinstance(query_info, str):
                    # Если это строка - SQL-запрос
                    query = query_info
                    params = None
                elif isinstance(query_info, tuple) and len(query_info) >= 1:
                    # Если это кортеж - (запрос, параметры)
                    query = query_info[0]
                    params = query_info[1] if len(query_info) > 1 else None
                else:
                    raise ValueError("Каждый запрос должен быть строкой или кортежем (запрос, параметры)")

                # Выполняем запрос и сохраняем результат
                result = self.execute_query(query, params)
                results.append(result)

            # Фиксируем транзакцию
            if not self.commit_transaction():
                raise Exception("Не удалось зафиксировать транзакцию")
            return results
        except Exception as e:
            # Если произошла ошибка - откатываем транзакцию
            try:
                self.rollback_transaction()
            except Exception as rollback_error:
                print(f"Ошибка при откате транзакции после основной ошибки: {rollback_error}")
            print(f"Транзакция не удалась и была откачена: {e}")
            raise e

    def export_to_excel(self, table_name, file_path, query=None):
        """
        Экспорт данных в Excel файл

        Args:
            table_name (str): Имя таблицы для экспорта
            file_path (str): Путь к файлу для сохранения
            query (str, optional): Пользовательский SQL-запрос для экспорта
        """
        try:
            if query:
                # Выполняем пользовательский запрос
                cursor = self.connection.cursor()
                cursor.execute(query)
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()

                # Используем первую часть запроса как название листа, если table_name пуст
                sheet_title = table_name if table_name else "query_result"
            else:
                # Получаем все данные из таблицы
                result = self.get_table_data(table_name)
                columns = result['columns']
                rows = []
                for row_dict in result['data']:
                    row = [row_dict.get(col) for col in columns]
                    rows.append(row)

                sheet_title = table_name

            # Создаем Excel файл
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = f"{sheet_title}_export"

            # Добавляем заголовки
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Добавляем данные
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # Сохраняем файл
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"Ошибка при экспорте в Excel: {e}")
            return False

    def export_to_csv(self, table_name, file_path, query=None):
        """
        Экспорт данных в CSV файл

        Args:
            table_name (str): Имя таблицы для экспорта
            file_path (str): Путь к файлу для сохранения
            query (str, optional): Пользовательский SQL-запрос для экспорта
        """
        import csv

        try:
            if query:
                # Выполняем пользовательский запрос
                cursor = self.connection.cursor()
                cursor.execute(query)
                columns = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
            else:
                # Получаем все данные из таблицы
                result = self.get_table_data(table_name)
                columns = result['columns']
                rows = []
                for row_dict in result['data']:
                    row = [row_dict.get(col) for col in columns]
                    rows.append(row)

            # Создаем CSV файл
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)

                # Записываем заголовки
                writer.writerow(columns)

                # Записываем данные
                writer.writerows(rows)

            return True
        except Exception as e:
            print(f"Ошибка при экспорте в CSV: {e}")
            return False


class TableModel(QAbstractTableModel):
    """
    Модель таблицы для отображения данных в QTableView
    
    Реализует интерфейс QAbstractTableModel для отображения 
    данных в табличном виде в PyQt6
    """
    
    def __init__(self, data):
        """
        Инициализация модели таблицы
        
        Args:
            data: Данные для отображения в таблице
        """
        super(TableModel, self).__init__()
        self._data = data if data else []  # Данные таблицы
        # Заголовки столбцов (по умолчанию числовые индексы)
        self._headers = [] if not data else [str(i) for i in range(len(data[0]))] if data else []

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        """
        Получение данных для отображения в ячейке
        
        Args:
            index: Индекс ячейки
            role: Роль данных (обычно DisplayRole для отображения текста)
            
        Returns:
            Данные для отображения в ячейке
        """
        if role == Qt.ItemDataRole.DisplayRole:
            return str(self._data[index.row()][index.column()])
        return None

    def rowCount(self, parent=None):
        """
        Возвращает количество строк в модели
        
        Args:
            parent: Родительский индекс (обычно None)
            
        Returns:
            int: Количество строк
        """
        return len(self._data)

    def columnCount(self, parent=None):
        """
        Возвращает количество столбцов в модели
        
        Args:
            parent: Родительский индекс (обычно None)
            
        Returns:
            int: Количество столбцов
        """
        return len(self._headers) if self._headers else (len(self._data[0]) if self._data else 0)

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        """
        Возвращает заголовки для строк или столбцов
        
        Args:
            section: Номер секции (строки или столбца)
            orientation: Ориентация (горизонтальная или вертикальная)
            role: Роль данных
            
        Returns:
            str: Заголовок
        """
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            return self._headers[section] if section < len(self._headers) else str(section)
        return super().headerData(section, orientation, role)