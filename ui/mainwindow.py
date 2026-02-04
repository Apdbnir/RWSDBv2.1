import sys
import os
from datetime import datetime
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView,
                             QPushButton, QComboBox, QLineEdit, QLabel, QGroupBox,
                             QFormLayout, QScrollArea, QSplitter, QFrame, QFileDialog,
                             QMessageBox, QDialog, QInputDialog, QCheckBox, QSpinBox, QTextEdit,
                             QAbstractItemView, QListWidget)
from PyQt6.QtCore import Qt, QAbstractTableModel, QModelIndex
from PyQt6.QtGui import QAction, QKeySequence
import openpyxl
from openpyxl.styles import Font, PatternFill

from database.manager import DatabaseManager, TableModel
from utils.backup.backup_manager import BackupManager
from utils.query_history import QueryHistoryManager
from utils.translations import TranslationManager
from utils.constants import *
from utils.security import *
from ui.dialogs.dialogs import AddEntityDialog, UpdateEntityDialog
from ui.sql_queries import (TrainQueries, TicketQueries, ServiceQueries, ScheduleQueries,
                           PlatformQueries, PassengerQueries, EmployeeQueries,
                           TrainTicketJoinQueries, ServiceEmployeeJoinQueries,
                           PlatformScheduleJoinQueries, EmployeeTrainJoinQueries,
                           ScheduleTrainJoinQueries, TicketPassengerJoinQueries,
                           TicketScheduleJoinQueries, AggregateQueries)
from ui.lab_queries import LabQueries
from ui.sql_syntax_highlighter import SqlSyntaxHighlighter
from utils.export_utils import ExportUtils
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import QStyle


class DatabaseApp(QMainWindow):
    def __init__(self, user_mode='admin'):
        super().__init__()
        self.user_mode = user_mode  # 'admin' or 'user'
        # Load admin password from configuration file or use default
        self.admin_password_hash, self.admin_salt = self.load_admin_password()
        self.db_manager = None
        self.query_history_manager = QueryHistoryManager()
        # Set up callback for automatic updates when queries are added/deleted
        self.query_history_manager.update_callback = self.refresh_saved_queries
        self.backup_manager = None
        # Initialize language before setting up UI
        self.translation_manager = TranslationManager()
        self.current_language = 'en'
        self.selected_language = 'en'

        # Apply application styling
        self.apply_styling()

        self.setup_ui()
        self.refresh_tables()

        # Set UI based on user mode
        self.set_ui_mode()

    def apply_styling(self):
        """Apply the application stylesheet"""
        try:
            import os
            style_file = os.path.join(os.path.dirname(__file__), 'styles.qss')
            if os.path.exists(style_file):
                with open(style_file, 'r') as f:
                    style = f.read()
                self.setStyleSheet(style)
        except Exception as e:
            print(f"Could not load stylesheet: {e}")

    def is_numeric(self, value):
        """Check if value is numeric (int or float)"""
        try:
            float(value)
            return True
        except ValueError:
            return False

    def is_date_or_time(self, value):
        """Check if value is a date or time format"""
        import re
        # Check various date/time formats
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
            r'\d{2}:\d{2}:\d{2}',  # HH:MM:SS
            r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}',  # YYYY-MM-DD HH:MM:SS
            r'\d{2}/\d{2}/\d{4}',  # MM/DD/YYYY
            r'\d{2}\.\d{2}\.\d{4}',  # DD.MM.YYYY
        ]

        for pattern in date_patterns:
            if re.match(pattern, value.strip()):
                return True
        return False

    def get_translation(self, key, language_code=None):
        """Get translated text based on key and language"""
        return self.translation_manager.get_translation(key, language_code)

    def set_ui_mode(self):
        """Configure UI based on user mode (admin/user)"""
        if self.user_mode == 'user':
            # Disable admin-only functionality
            self.add_row_btn.setEnabled(False)
            self.delete_row_btn.setEnabled(False)
            self.edit_row_btn.setEnabled(False)
            self.execute_btn.setEnabled(False)  # Disable direct query execution that modifies data
            self.execute_transaction_btn.setEnabled(False)
            self.add_table_btn.setEnabled(False)
            self.delete_table_btn.setEnabled(False)
            self.add_column_btn.setEnabled(False)
            self.delete_column_btn.setEnabled(False)
            self.edit_column_btn.setEnabled(False)

            # Disable menu items that allow data modification
            self.find_admin_menu_items()

            self.statusBar().showMessage(self.get_translation("User mode: Read-only access"))
        else:
            self.statusBar().showMessage(self.get_translation("Administrator mode: Full access"))

    def find_admin_menu_items(self):
        """Find and disable menu items that should be restricted in user mode"""
        # This is a placeholder - in a more complex implementation you would track menu items
        pass

    def app(self):
        """Helper method to get the application instance"""
        from PyQt6.QtWidgets import QApplication
        return QApplication.instance()

    def setup_ui(self):
        self.setWindowTitle(self.get_translation(APP_NAME))
        self.setGeometry(100, 100, WINDOW_WIDTH, WINDOW_HEIGHT)
        self.setMinimumSize(MIN_WINDOW_WIDTH, MIN_WINDOW_HEIGHT)

        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Create menu bar
        self.create_menu_bar()
        # Store references to menu items that need to be translated when language changes
        # Now translate the menu items to the current language

        # Create main splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # Left panel for database navigation
        left_panel = self.create_left_panel()
        splitter.addWidget(left_panel)

        # Right panel for table view and query execution
        right_panel = self.create_right_panel()
        splitter.addWidget(right_panel)

        splitter.setSizes([300, 1100])
        main_layout.addWidget(splitter)

        # Create toolbar for undo/redo
        self.create_toolbar()

        # Status bar
        self.statusBar().showMessage(self.get_translation("Ready"))

    def create_toolbar(self):
        """Create toolbar with undo/redo buttons"""
        toolbar = self.addToolBar('UndoRedo')

        # Create undo action
        self.undo_action = QAction('Undo', self)
        self.undo_action.setShortcut(QKeySequence.StandardKey.Undo)
        self.undo_action.triggered.connect(self.undo_action_triggered)
        self.undo_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowBack))
        self.undo_action.setEnabled(False)  # Initially disabled
        # Set object name for CSS styling (change/modify for undo)
        undo_btn = toolbar.widgetForAction(self.undo_action)
        if undo_btn:
            undo_btn.setObjectName("change")
        toolbar.addAction(self.undo_action)

        # Create redo action
        self.redo_action = QAction('Redo', self)
        self.redo_action.setShortcut(QKeySequence.StandardKey.Redo)
        self.redo_action.triggered.connect(self.redo_action_triggered)
        self.redo_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowForward))
        self.redo_action.setEnabled(False)  # Initially disabled
        # Set object name for CSS styling (change/modify for redo)
        redo_btn = toolbar.widgetForAction(self.redo_action)
        if redo_btn:
            redo_btn.setObjectName("change")
        toolbar.addAction(self.redo_action)

        # Update button states
        self.update_undo_redo_states()

    def create_menu_bar(self):
        menubar = self.menuBar()

        # File menu
        file_menu = menubar.addMenu(self.get_translation('File'))

        connect_action = QAction(self.get_translation('Connect to Database'), self)
        connect_action.triggered.connect(self.connect_to_database)
        connect_action.setToolTip("Connect to a database file")
        from PyQt6.QtWidgets import QStyle
        connect_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DirIcon))
        file_menu.addAction(connect_action)

        backup_action = QAction(self.get_translation('Create Backup'), self)
        backup_action.triggered.connect(self.create_backup)
        backup_action.setToolTip("Create a backup of the current database")
        backup_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        file_menu.addAction(backup_action)

        restore_action = QAction(self.get_translation('Restore from Backup'), self)
        restore_action.triggered.connect(self.restore_backup)
        restore_action.setToolTip("Restore the database from a backup file")
        restore_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogOpenButton))
        file_menu.addAction(restore_action)

        export_action = QAction(self.get_translation('Export to...'), self)
        export_action.triggered.connect(self.export_table_to_excel)
        export_action.setToolTip("Export the selected table to various formats (Excel, Text, Markdown, SQL)")
        export_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton))
        file_menu.addAction(export_action)

        # Table menu
        table_menu = menubar.addMenu(self.get_translation('Table'))

        add_table_action = QAction(self.get_translation('Add New Table'), self)
        add_table_action.triggered.connect(self.add_new_table)
        add_table_action.setToolTip("Add a new table to the database")
        add_table_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileIcon))
        table_menu.addAction(add_table_action)

        delete_table_action = QAction(self.get_translation('Delete Table'), self)
        delete_table_action.triggered.connect(self.delete_table)
        delete_table_action.setToolTip("Delete the selected table from the database")
        delete_table_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_TrashIcon))
        table_menu.addAction(delete_table_action)

        # Column menu
        column_menu = menubar.addMenu(self.get_translation('Column'))

        add_column_action = QAction(self.get_translation('Add Column'), self)
        add_column_action.triggered.connect(self.add_column)
        add_column_action.setToolTip("Add a new column to the selected table")
        add_column_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileIcon))
        column_menu.addAction(add_column_action)

        delete_column_action = QAction(self.get_translation('Delete Column'), self)
        delete_column_action.triggered.connect(self.delete_column)
        delete_column_action.setToolTip("Delete the selected column from the table")
        delete_column_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_TrashIcon))
        column_menu.addAction(delete_column_action)

        edit_column_action = QAction(self.get_translation('Edit Column'), self)
        edit_column_action.triggered.connect(self.edit_column)
        edit_column_action.setToolTip("Edit the selected column (supports PostgreSQL ALTER TABLE operations)")
        edit_column_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogDetailedView))
        column_menu.addAction(edit_column_action)

        # Special Queries menu
        special_queries_menu = menubar.addMenu(self.get_translation('Special Queries'))

        # Add predefined queries for the train station database
        query1_action = QAction(self.get_translation('get_all_trains'), self)
        query1_action.triggered.connect(lambda: self.execute_predefined_query("SELECT * FROM train"))
        query1_action.setToolTip("Get all train records with their details")
        query1_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query1_action)

        query2_action = QAction(self.get_translation('get_all_employees'), self)
        query2_action.triggered.connect(lambda: self.execute_predefined_query("SELECT full_name, position FROM employee ORDER BY position"))
        query2_action.setToolTip("Get all employees with their positions ordered by position")
        query2_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query2_action)

        query3_action = QAction(self.get_translation('get_scheduled_arrivals'), self)
        query3_action.triggered.connect(lambda: self.execute_predefined_query("SELECT arrival_time, departure_time, train_id, platform_id FROM schedule"))
        query3_action.setToolTip("Get scheduled arrivals and departures")
        query3_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query3_action)

        query4_action = QAction(self.get_translation('get_passengers_with_tickets'), self)
        query4_action.triggered.connect(lambda: self.execute_predefined_query("SELECT p.full_name, p.passport_number, t.ticket_id, t.seat_number FROM passenger p JOIN ticket t ON p.passenger_id = t.passenger_id"))
        query4_action.setToolTip("Get passengers with their ticket information")
        query4_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query4_action)

        query5_action = QAction(self.get_translation('get_employee_train_assignments'), self)
        query5_action.triggered.connect(lambda: self.execute_predefined_query("SELECT e.full_name, e.position, t.train_id FROM employee e JOIN assignment a ON e.employee_id = a.employee_id JOIN train t ON a.train_id = t.train_id"))
        query5_action.setToolTip("Get assignment of employees to trains")
        query5_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query5_action)

        query6_action = QAction(self.get_translation('get_all_scheduled_trains'), self)
        query6_action.triggered.connect(lambda: self.execute_predefined_query("SELECT s.schedule_id, s.arrival_time, s.departure_time, s.date, t.train_id, tr.type, p.platform_id, p.location, p.capacity FROM schedule s JOIN train tr ON s.train_id = tr.train_id JOIN platform p ON s.platform_id = p.platform_id"))
        query6_action.setToolTip("Get all scheduled trains with platform information")
        query6_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query6_action)

        query7_action = QAction(self.get_translation('get_passengers_tickets_trains'), self)
        query7_action.triggered.connect(lambda: self.execute_predefined_query("SELECT pass.full_name AS passenger_name, pass.passport_number, t.ticket_id, t.seat_number, tr.train_id, tr.type FROM passenger pass JOIN ticket t ON pass.passenger_id = t.passenger_id JOIN schedule s ON t.ticket_id = s.ticket_id JOIN train tr ON s.train_id = tr.train_id"))
        query7_action.setToolTip("Get passengers with their tickets and train information")
        query7_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query7_action)

        query8_action = QAction(self.get_translation('get_employees_train_details'), self)
        query8_action.triggered.connect(lambda: self.execute_predefined_query("SELECT e.employee_id, e.full_name, e.position, e.experience, t.train_id, t.type FROM employee e JOIN assignment a ON e.employee_id = a.employee_id JOIN train t ON a.train_id = t.train_id"))
        query8_action.setToolTip("Get employees assigned to specific trains with their details")
        query8_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query8_action)

        query9_action = QAction(self.get_translation('get_services_employee_assignments'), self)
        query9_action.triggered.connect(lambda: self.execute_predefined_query("SELECT s.service_name, s.price, s.type, s.service_date, e.full_name AS employee_name, e.position FROM service s JOIN service_assignment sa ON s.service_id = sa.service_id JOIN employee e ON sa.employee_id = e.employee_id"))
        query9_action.setToolTip("Get services with assignments to employees")
        query9_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query9_action)

        query10_action = QAction(self.get_translation('get_detailed_schedule_passengers'), self)
        query10_action.triggered.connect(lambda: self.execute_predefined_query("SELECT s.arrival_time, s.departure_time, s.date, tr.train_id, tr.type, p.platform_id, pass.full_name AS passenger_name, t.seat_number FROM schedule s JOIN train tr ON s.train_id = tr.train_id JOIN platform p ON s.platform_id = p.platform_id JOIN ticket t ON s.ticket_id = t.ticket_id JOIN passenger pass ON t.passenger_id = pass.passenger_id"))
        query10_action.setToolTip("Get detailed schedule with passenger information")
        query10_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query10_action)

        # Additional queries based on train station operations
        query11_action = QAction(self.get_translation("get_all_passengers_by_train"), self)
        query11_action.triggered.connect(lambda: self.execute_predefined_query("SELECT tr.type, tr.train_id, pass.full_name, t.seat_number FROM train tr JOIN schedule s ON tr.train_id = s.train_id JOIN ticket t ON s.ticket_id = t.ticket_id JOIN passenger pass ON t.passenger_id = pass.passenger_id ORDER BY tr.train_id"))
        query11_action.setToolTip("Get all passengers by train")
        query11_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query11_action)

        query12_action = QAction(self.get_translation("get_employee_assignments_by_train_type"), self)
        query12_action.triggered.connect(lambda: self.execute_predefined_query("SELECT tr.type, e.full_name, e.position FROM train tr JOIN assignment a ON tr.train_id = a.train_id JOIN employee e ON a.employee_id = e.employee_id ORDER BY tr.type"))
        query12_action.setToolTip("Get employee assignments by train type")
        query12_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query12_action)

        query13_action = QAction(self.get_translation("get_platform_utilization"), self)
        query13_action.triggered.connect(lambda: self.execute_predefined_query("SELECT p.location, p.capacity, COUNT(s.schedule_id) as scheduled_trains FROM platform p LEFT JOIN schedule s ON p.platform_id = s.platform_id GROUP BY p.platform_id, p.location, p.capacity ORDER BY p.platform_id"))
        query13_action.setToolTip("Get platform utilization statistics")
        query13_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query13_action)

        query14_action = QAction(self.get_translation("get_tickets_by_price_range"), self)
        query14_action.triggered.connect(lambda: self.execute_predefined_query("SELECT t.ticket_id, t.seat_number, t.price, pass.full_name FROM ticket t JOIN passenger pass ON t.passenger_id = pass.passenger_id WHERE t.price BETWEEN 10 AND 50 ORDER BY t.price"))
        query14_action.setToolTip("Get tickets within a specific price range")
        query14_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query14_action)

        query15_action = QAction(self.get_translation("get_scheduled_trains_by_date"), self)
        query15_action.triggered.connect(lambda: self.execute_predefined_query("SELECT s.date, s.arrival_time, s.departure_time, tr.type, tr.train_id, p.location FROM schedule s JOIN train tr ON s.train_id = tr.train_id JOIN platform p ON s.platform_id = p.platform_id WHERE s.date = '2023-12-01' ORDER BY s.arrival_time"))
        query15_action.setToolTip("Get scheduled trains for a specific date")
        query15_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_ComputerIcon))
        special_queries_menu.addAction(query15_action)

        # Transaction menu
        transaction_menu = menubar.addMenu(self.get_translation('Transaction'))

        begin_transaction_action = QAction(self.get_translation('Begin Transaction'), self)
        begin_transaction_action.triggered.connect(self.begin_transaction)
        begin_transaction_action.setToolTip("Begin a new database transaction")
        begin_transaction_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaPlay))
        transaction_menu.addAction(begin_transaction_action)

        commit_transaction_action = QAction(self.get_translation('Commit Transaction'), self)
        commit_transaction_action.triggered.connect(self.commit_transaction)
        commit_transaction_action.setToolTip("Commit the current database transaction")
        commit_transaction_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton))
        transaction_menu.addAction(commit_transaction_action)

        rollback_transaction_action = QAction(self.get_translation('Rollback Transaction'), self)
        rollback_transaction_action.triggered.connect(self.rollback_transaction)
        rollback_transaction_action.setToolTip("Rollback the current database transaction")
        rollback_transaction_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_BrowserReload))
        transaction_menu.addAction(rollback_transaction_action)


        # Settings menu
        settings_menu = menubar.addMenu(self.get_translation('Settings'))

        # Password change action (only available in admin mode)
        if self.user_mode == 'admin':
            change_password_action = QAction(self.get_translation('Change Password'), self)
            change_password_action.triggered.connect(self.change_password)
            change_password_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogContentsView))
            settings_menu.addAction(change_password_action)

        # Language submenu
        language_menu = settings_menu.addMenu(self.get_translation('Language'))
        language_menu.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DesktopIcon))

        self.lang_ru_action = QAction(self.get_translation('Русский'), self)
        self.lang_ru_action.setCheckable(True)
        self.lang_ru_action.triggered.connect(lambda: self.change_language('ru'))
        self.lang_ru_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DesktopIcon))
        language_menu.addAction(self.lang_ru_action)

        self.lang_be_action = QAction(self.get_translation('Беларускі'), self)
        self.lang_be_action.setCheckable(True)
        self.lang_be_action.triggered.connect(lambda: self.change_language('be'))
        self.lang_be_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DesktopIcon))
        language_menu.addAction(self.lang_be_action)

        self.lang_en_action = QAction(self.get_translation('English'), self)
        self.lang_en_action.setCheckable(True)
        self.lang_en_action.triggered.connect(lambda: self.change_language('en'))
        self.lang_en_action.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DesktopIcon))
        language_menu.addAction(self.lang_en_action)

        # Set default language as English
        self.lang_en_action.setChecked(True)

    def change_password(self):
        """Open the password change dialog"""
        from ui.password_change_dialog import PasswordChangeDialog

        # Pass the current password hash and salt to the dialog
        dialog = PasswordChangeDialog(self.admin_password_hash, self.admin_salt)
        if dialog.exec():
            new_password = dialog.get_new_password()
            if new_password:
                # Save the new password to the configuration file
                if self.save_admin_password(new_password):
                    # Update the local variables
                    pwd_hash, salt = hash_password(new_password)
                    self.admin_password_hash = pwd_hash
                    self.admin_salt = salt
                    QMessageBox.information(
                        self,
                        "Password Changed",
                        "Administrator password has been successfully changed."
                    )
                    self.statusBar().showMessage(self.get_translation("Password changed successfully"))
                else:
                    QMessageBox.warning(
                        self,
                        "Error",
                        "Password change failed."
                    )
            else:
                QMessageBox.warning(
                    self,
                    "Error",
                    "Password change failed."
                )

    def load_admin_password(self):
        """Load the admin password hash and salt from a configuration file"""
        config = load_secure_config('config.json')
        # Return the admin password hash and salt from config
        pwd_hash = config.get('admin_password_hash', '')
        salt_hex = config.get('admin_salt', '')
        salt = bytes.fromhex(salt_hex) if salt_hex else b''
        return pwd_hash, salt

    def save_admin_password(self, password):
        """Save the admin password hash and salt to a configuration file"""
        # Hash the password and get the salt
        pwd_hash, salt = hash_password(password)

        # Load existing config
        config = load_secure_config('config.json')

        # Update the password hash and salt in config
        config['admin_password_hash'] = pwd_hash
        config['admin_salt'] = salt.hex()

        # Save the updated config
        return save_secure_config(config, 'config.json')

    def begin_transaction(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        try:
            if self.db_manager.begin_transaction():
                self.statusBar().showMessage(self.get_translation("transaction_started_successfully"))
                QMessageBox.information(self, "Transaction", self.get_translation("Transaction started successfully. All subsequent operations will be part of this transaction until committed or rolled back."))
            else:
                QMessageBox.critical(self, "Error", self.get_translation("Failed to begin transaction."))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error beginning transaction: {str(e)}")

    def commit_transaction(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        reply = QMessageBox.question(
            self,
            "Confirm Commit",
            self.get_translation("Are you sure you want to commit the current transaction? This will make all changes permanent."),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                if self.db_manager.commit_transaction():
                    self.statusBar().showMessage(self.get_translation("transaction_committed_successfully"))
                    QMessageBox.information(self, "Transaction", self.get_translation("Transaction committed successfully. All changes are now permanent."))
                else:
                    QMessageBox.critical(self, "Error", self.get_translation("Failed to commit transaction."))
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error committing transaction: {str(e)}")

    def rollback_transaction(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        reply = QMessageBox.question(
            self,
            "Confirm Rollback",
            self.get_translation("Are you sure you want to rollback the current transaction? This will undo all changes made since the transaction began."),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                if self.db_manager.rollback_transaction():
                    self.statusBar().showMessage(self.get_translation("transaction_rolled_back_successfully"))
                    QMessageBox.information(self, "Transaction", self.get_translation("Transaction rolled back successfully. All changes have been undone."))
                else:
                    QMessageBox.critical(self, "Error", self.get_translation("Failed to rollback transaction."))
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error rolling back transaction: {str(e)}")


    def create_left_panel(self):
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)

        # Tables group box
        tables_group = QGroupBox(self.get_translation("Table"))
        tables_layout = QVBoxLayout(tables_group)

        # Tables list
        self.tables_list = QTableWidget()
        self.tables_list.setColumnCount(1)
        self.tables_list.setHorizontalHeaderLabels([self.get_translation("Table")])
        self.tables_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.tables_list.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)  # Make tables list read-only
        self.tables_list.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tables_list.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.tables_list.cellClicked.connect(self.table_selected)
        self.tables_list.setToolTip("List of database tables")
        tables_layout.addWidget(self.tables_list)

        # Add table button
        self.add_table_btn = QPushButton(self.get_translation("Add New Table"))
        self.add_table_btn.setObjectName("add")  # Green on hover for add
        self.add_table_btn.clicked.connect(self.add_new_table)
        self.add_table_btn.setToolTip("Add a new table to the database")
        tables_layout.addWidget(self.add_table_btn)

        # Delete table button
        self.delete_table_btn = QPushButton(self.get_translation("Delete Table"))
        self.delete_table_btn.setObjectName("delete")  # Red on hover for delete
        self.delete_table_btn.clicked.connect(self.delete_table)
        self.delete_table_btn.setToolTip("Delete the selected table from the database")
        tables_layout.addWidget(self.delete_table_btn)

        # Refresh button
        self.refresh_btn = QPushButton(self.get_translation("Refresh Data"))
        self.refresh_btn.setObjectName("refresh")  # Blue on hover for refresh
        self.refresh_btn.clicked.connect(self.refresh_tables)
        self.refresh_btn.setToolTip("Refresh the list of tables and their data")
        tables_layout.addWidget(self.refresh_btn)

        left_layout.addWidget(tables_group)

        # Columns group box
        columns_group = QGroupBox(self.get_translation("Column"))
        columns_layout = QVBoxLayout(columns_group)

        # Columns list
        self.columns_list = QTableWidget()
        self.columns_list.setColumnCount(5)
        self.columns_list.setHorizontalHeaderLabels([self.get_translation("Name"), self.get_translation("Type"), self.get_translation("Not Null"), self.get_translation("Default"), self.get_translation("PK")])
        self.columns_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.columns_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.columns_list.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.columns_list.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.columns_list.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.columns_list.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)  # Make column list read-only
        self.columns_list.setToolTip("List of columns for the selected table")
        columns_layout.addWidget(self.columns_list)

        # Add column button
        self.add_column_btn = QPushButton(self.get_translation("Add Column"))
        self.add_column_btn.setObjectName("add")  # Green on hover for add
        self.add_column_btn.clicked.connect(self.add_column)
        self.add_column_btn.setToolTip("Add a new column to the selected table")
        columns_layout.addWidget(self.add_column_btn)

        # Delete column button
        self.delete_column_btn = QPushButton(self.get_translation("Delete Column"))
        self.delete_column_btn.setObjectName("delete")  # Red on hover for delete
        self.delete_column_btn.clicked.connect(self.delete_column)
        self.delete_column_btn.setToolTip("Delete the selected column from the table")
        columns_layout.addWidget(self.delete_column_btn)

        # Edit column button
        self.edit_column_btn = QPushButton(self.get_translation("Edit Column"))
        self.edit_column_btn.setObjectName("edit")  # Blue on hover for edit
        self.edit_column_btn.clicked.connect(self.edit_column)
        self.edit_column_btn.setToolTip("Edit the selected column (supports PostgreSQL ALTER TABLE operations)")
        columns_layout.addWidget(self.edit_column_btn)

        left_layout.addWidget(columns_group)

        # Make the left panel scrollable
        scroll_area = QScrollArea()
        scroll_area.setWidget(left_widget)
        scroll_area.setWidgetResizable(True)

        return scroll_area

    def create_right_panel(self):
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)

        # Create tab widget
        self.tab_widget = QTabWidget()
        self.tab_widget.setToolTip("Tabbed interface for different database operations")

        # Table View tab
        self.table_view_tab = self.create_table_view_tab()
        self.tab_widget.addTab(self.table_view_tab, self.get_translation("table_view_tab"))

        # Special Queries tab
        self.special_queries_tab = self.create_special_queries_tab()
        self.tab_widget.addTab(self.special_queries_tab, self.get_translation("special_queries_tab"))

        # Query Editor tab
        self.query_tab = self.create_query_tab()
        self.tab_widget.addTab(self.query_tab, self.get_translation("query_editor_tab"))

        # Query History tab
        self.history_tab = self.create_history_tab()
        self.tab_widget.addTab(self.history_tab, self.get_translation("query_history_tab"))



        # Saved Queries tab
        self.saved_queries_tab = self.create_saved_queries_tab()
        self.tab_widget.addTab(self.saved_queries_tab, "Saved Queries")

        right_layout.addWidget(self.tab_widget)

        return right_widget

    def create_table_view_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Table view
        self.table_view = QTableWidget()
        self.table_view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)  # Make table view read-only to enforce validation through dialogs
        self.table_view.setToolTip("Display data from the selected table")
        layout.addWidget(self.table_view)

        # Table controls
        controls_layout = QHBoxLayout()

        self.refresh_data_btn = QPushButton(self.get_translation("Refresh Data"))
        self.refresh_data_btn.setObjectName("blue")  # Blue for refresh
        self.refresh_data_btn.clicked.connect(self.refresh_current_table)
        self.refresh_data_btn.setToolTip("Refresh data in the current table view")
        controls_layout.addWidget(self.refresh_data_btn)

        self.add_row_btn = QPushButton(self.get_translation("Add Row"))
        self.add_row_btn.setObjectName("add")  # Green on hover for add
        self.add_row_btn.clicked.connect(self.add_row)
        self.add_row_btn.setToolTip("Add a new row to the selected table")
        controls_layout.addWidget(self.add_row_btn)

        self.delete_row_btn = QPushButton(self.get_translation("Delete Row"))
        self.delete_row_btn.setObjectName("delete")  # Red on hover for delete
        self.delete_row_btn.clicked.connect(self.delete_row)
        self.delete_row_btn.setToolTip("Delete the selected row from the table")
        controls_layout.addWidget(self.delete_row_btn)

        self.edit_row_btn = QPushButton(self.get_translation("Edit Row"))
        self.edit_row_btn.setObjectName("edit")  # Blue on hover for edit
        self.edit_row_btn.clicked.connect(self.edit_row)
        self.edit_row_btn.setToolTip("Edit the selected row (not implemented in this version)")
        controls_layout.addWidget(self.edit_row_btn)

        layout.addLayout(controls_layout)

        # Filter controls
        filter_layout = QHBoxLayout()

        filter_label = QLabel("Filter:")
        filter_layout.addWidget(filter_label)

        self.filter_column_combo = QComboBox()
        self.filter_column_combo.setToolTip("Select column to filter by")
        filter_layout.addWidget(self.filter_column_combo)

        self.filter_value_input = QLineEdit()
        self.filter_value_input.setPlaceholderText("Enter filter value...")
        self.filter_value_input.setToolTip("Enter value to filter by")
        filter_layout.addWidget(self.filter_value_input)

        self.apply_filter_btn = QPushButton("Apply Filter")
        self.apply_filter_btn.setObjectName("blue")
        self.apply_filter_btn.clicked.connect(self.apply_table_filter)
        self.apply_filter_btn.setToolTip("Apply filter to current table and show results in Query Editor")
        filter_layout.addWidget(self.apply_filter_btn)

        self.clear_filter_btn = QPushButton("Clear Filter")
        self.clear_filter_btn.setObjectName("cancel")
        self.clear_filter_btn.clicked.connect(self.clear_table_filter)
        self.clear_filter_btn.setToolTip("Clear current filter")
        filter_layout.addWidget(self.clear_filter_btn)

        layout.addLayout(filter_layout)

        # Pagination controls
        pagination_layout = QHBoxLayout()

        pagination_label = QLabel("Page:")
        pagination_layout.addWidget(pagination_label)

        self.page_spinbox = QSpinBox()
        self.page_spinbox.setRange(1, 1)
        self.page_spinbox.setValue(1)
        self.page_spinbox.setToolTip("Current page number")
        pagination_layout.addWidget(self.page_spinbox)

        self.total_pages_label = QLabel("/ 1")
        self.total_pages_label.setToolTip("Total number of pages")
        pagination_layout.addWidget(self.total_pages_label)

        self.rows_per_page_combo = QComboBox()
        self.rows_per_page_combo.addItems(["100", "500", "1000", "5000"])
        self.rows_per_page_combo.setCurrentText("1000")
        self.rows_per_page_combo.setToolTip("Number of rows per page")
        pagination_layout.addWidget(self.rows_per_page_combo)

        self.first_page_btn = QPushButton("|<<")
        self.first_page_btn.setObjectName("blue")
        self.first_page_btn.clicked.connect(self.go_to_first_page)
        self.first_page_btn.setToolTip("Go to first page")
        pagination_layout.addWidget(self.first_page_btn)

        self.prev_page_btn = QPushButton("<<")
        self.prev_page_btn.setObjectName("blue")
        self.prev_page_btn.clicked.connect(self.go_to_prev_page)
        self.prev_page_btn.setToolTip("Go to previous page")
        pagination_layout.addWidget(self.prev_page_btn)

        self.next_page_btn = QPushButton(">>")
        self.next_page_btn.setObjectName("blue")
        self.next_page_btn.clicked.connect(self.go_to_next_page)
        self.next_page_btn.setToolTip("Go to next page")
        pagination_layout.addWidget(self.next_page_btn)

        self.last_page_btn = QPushButton(">>|")
        self.last_page_btn.setObjectName("blue")
        self.last_page_btn.clicked.connect(self.go_to_last_page)
        self.last_page_btn.setToolTip("Go to last page")
        pagination_layout.addWidget(self.last_page_btn)

        # Add stretch to push pagination controls to the right
        pagination_layout.addStretch()

        layout.addLayout(pagination_layout)

        return widget

    def create_query_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Query input
        self.query_input = QTextEdit()
        self.query_input.setPlaceholderText(self.get_translation("Enter SQL query here..."))

        # Initialize SQL syntax highlighter
        self.sql_highlighter = SqlSyntaxHighlighter(self.query_input.document())

        layout.addWidget(self.query_input)

        # Query controls
        query_controls = QHBoxLayout()

        self.execute_btn = QPushButton(self.get_translation("Execute Query"))
        self.execute_btn.setObjectName("ok")  # Green on hover for execute
        self.execute_btn.clicked.connect(self.execute_query)
        self.execute_btn.setToolTip("Execute the SQL query in the input field")
        query_controls.addWidget(self.execute_btn)

        self.execute_transaction_btn = QPushButton(self.get_translation("Execute as Transaction"))
        self.execute_transaction_btn.setObjectName("ok")  # Green on hover for transaction
        self.execute_transaction_btn.clicked.connect(self.execute_transaction_query)
        self.execute_transaction_btn.setToolTip("Execute multiple queries as a single transaction")
        query_controls.addWidget(self.execute_transaction_btn)

        self.save_query_btn = QPushButton(self.get_translation("Save Query"))
        self.save_query_btn.setObjectName("save")  # Green on hover for save
        self.save_query_btn.clicked.connect(self.save_query)
        self.save_query_btn.setToolTip("Save the current query to query history")
        query_controls.addWidget(self.save_query_btn)

        self.export_query_btn = QPushButton(self.get_translation("Export to..."))
        self.export_query_btn.setObjectName("save")  # Green on hover for export
        self.export_query_btn.clicked.connect(self.export_query_results)
        self.export_query_btn.setToolTip("Export query results to various formats (Excel, Text, Markdown, SQL)")
        query_controls.addWidget(self.export_query_btn)

        layout.addLayout(query_controls)

        # Query results
        self.query_results = QTableWidget()
        self.query_results.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)  # Make query results read-only
        self.query_results.setToolTip("Display results of executed SQL queries")
        layout.addWidget(self.query_results)

        return widget

    def create_history_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # History list
        self.history_list = QTableWidget()
        self.history_list.setColumnCount(3)
        self.history_list.setHorizontalHeaderLabels([self.get_translation("Description"), self.get_translation("Query"), self.get_translation("Timestamp")])
        self.history_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.history_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.history_list.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.history_list.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)  # Make history list read-only
        self.history_list.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.history_list.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.history_list.cellDoubleClicked.connect(self.load_query_from_history)
        self.history_list.setToolTip("List of previously executed queries")
        layout.addWidget(self.history_list)

        # History controls
        history_controls = QHBoxLayout()

        self.load_query_btn = QPushButton(self.get_translation("Load Query"))
        self.load_query_btn.setObjectName("ok")  # Green on hover for load
        self.load_query_btn.clicked.connect(self.load_selected_query)
        self.load_query_btn.setToolTip("Load the selected query from history to the query editor")
        history_controls.addWidget(self.load_query_btn)

        self.delete_history_btn = QPushButton(self.get_translation("Delete Query"))
        self.delete_history_btn.setObjectName("delete")  # Red on hover for delete
        self.delete_history_btn.clicked.connect(self.delete_query_from_history)
        self.delete_history_btn.setToolTip("Delete the selected query from history")
        history_controls.addWidget(self.delete_history_btn)

        layout.addLayout(history_controls)

        self.refresh_query_history()

        return widget



    def create_select_queries_tab(self):
        """Create a tab for SELECT queries (single tables)"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Instructions label
        instructions = QLabel(self.get_translation("Select and execute SELECT queries for individual tables:"))
        instructions.setWordWrap(True)
        layout.addWidget(instructions)

        # Controls layout at the top (only selection controls, no execute button)
        controls_layout = QHBoxLayout()

        # Table selection for SELECT queries
        table_label = QLabel(self.get_translation("select_table_for_select_queries") + ":")
        controls_layout.addWidget(table_label)

        self.select_table_combo = QComboBox()
        self.select_table_combo.addItems([
            "Train", "Ticket", "Service", "Schedule", "Platform", "Passenger", "Employee"
        ])
        self.select_table_combo.currentTextChanged.connect(self.update_select_queries_display)
        controls_layout.addWidget(self.select_table_combo)

        # Query type selection
        query_type_label = QLabel(self.get_translation("select_query_type") + ":")
        controls_layout.addWidget(query_type_label)

        self.select_query_type_combo = QComboBox()
        self.select_query_type_combo.addItems([
            "SELECT FROM",
            "SELECT FROM WHERE",
            "SELECT FROM ORDER BY",
            "Combined Query"
        ])
        self.select_query_type_combo.currentTextChanged.connect(self.update_select_queries_display)
        controls_layout.addWidget(self.select_query_type_combo)

        # Add stretch to push controls to the left
        controls_layout.addStretch()

        layout.addLayout(controls_layout)

        # Create scroll area for query selection
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # Store reference to the scroll layout to update dynamically
        self.select_queries_scroll_layout = scroll_layout

        # Initial population of queries
        self.update_select_queries_display()

        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        # Action buttons at the bottom
        action_layout = QHBoxLayout()

        # Save query button only (execute button removed per user request)
        self.save_select_query_btn = QPushButton(self.get_translation("save_query"))
        self.save_select_query_btn.setObjectName("save")  # Green on hover for save
        self.save_select_query_btn.clicked.connect(self.save_select_query)
        action_layout.addWidget(self.save_select_query_btn)

        # Add stretch to push buttons to the left
        action_layout.addStretch()

        layout.addLayout(action_layout)

        return widget
        passenger_where_btn = QPushButton("Passenger: WHERE ID 5-10")
        passenger_where_btn.clicked.connect(lambda: self.execute_specific_query(PassengerQueries.SELECT_ID_BETWEEN_5_AND_10))
        passenger_layout.addWidget(passenger_from_btn)
        passenger_layout.addWidget(passenger_where_btn)
        scroll_layout.addLayout(passenger_layout)

        passenger_order_btn = QPushButton("Passenger: ORDER BY name DESC")
        passenger_order_btn.clicked.connect(lambda: self.execute_specific_query(PassengerQueries.SELECT_ORDER_BY_NAME_REVERSE))
        passenger_combined_btn = QPushButton("Passenger: Combined Query")
        passenger_combined_btn.clicked.connect(lambda: self.execute_specific_query(PassengerQueries.SELECT_COMBINED_PASSENGER))
        passenger2_layout = QHBoxLayout()
        passenger2_layout.addWidget(passenger_order_btn)
        passenger2_layout.addWidget(passenger_combined_btn)
        scroll_layout.addLayout(passenger2_layout)

        # Employee queries
        employee_layout = QHBoxLayout()
        employee_from_btn = QPushButton("Employee: SELECT first 5 with exp > 10")
        employee_from_btn.clicked.connect(lambda: self.execute_specific_query(EmployeeQueries.SELECT_FIRST_5_WITH_EXPERIENCE_10))
        employee_where_btn = QPushButton("Employee: WHERE position = engineer")
        employee_where_btn.clicked.connect(lambda: self.execute_specific_query(EmployeeQueries.SELECT_ONLY_ENGINEERS))
        employee_layout.addWidget(employee_from_btn)
        employee_layout.addWidget(employee_where_btn)
        scroll_layout.addLayout(employee_layout)

        employee_order_btn = QPushButton("Employee: ORDER BY passport DESC")
        employee_order_btn.clicked.connect(lambda: self.execute_specific_query(EmployeeQueries.SELECT_ORDER_BY_PASSPORT_REVERSE))
        employee_combined_btn = QPushButton("Employee: Combined Query")
        employee_combined_btn.clicked.connect(lambda: self.execute_specific_query(EmployeeQueries.SELECT_COMBINED_EMPLOYEE))
        employee2_layout = QHBoxLayout()
        employee2_layout.addWidget(employee_order_btn)
        employee2_layout.addWidget(employee_combined_btn)
        scroll_layout.addLayout(employee2_layout)

        # Add some stretch to push content to top
        scroll_layout.addStretch()

        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        return widget

    def update_select_queries_display(self):
        """Update the displayed queries based on selected table and query type"""
        # Clear existing widgets from the layout
        while self.select_queries_scroll_layout.count():
            child = self.select_queries_scroll_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
            elif child.layout():
                self.clear_layout(child.layout())

        # Get selected values
        table = self.select_table_combo.currentText()
        query_type = self.select_query_type_combo.currentText()

        if not table:
            # If no table is selected, show a message
            no_selection_label = QLabel("Please select a table to see query options")
            self.select_queries_scroll_layout.addWidget(no_selection_label)
            self.select_queries_scroll_layout.addStretch()
            return

        # Dynamically create query button based on selection
        query_button = QPushButton(f"{table}: {query_type}")
        query_button.setObjectName("ok")  # Green for execute

        # Generate generic queries based on table and query type since we can't rely on pre-defined queries
        generic_query = ""

        if query_type == "SELECT FROM":
            generic_query = f"SELECT * FROM {table}"
        elif query_type == "SELECT FROM WHERE":
            # Try to get table info to suggest a proper WHERE clause
            try:
                if self.db_manager:
                    table_info = self.db_manager.get_table_info(table)
                    # Find first non-ID column to use in WHERE clause
                    for col in table_info:
                        if col['name'].lower() != 'id':
                            generic_query = f"SELECT * FROM {table} WHERE {col['name']} = ?"
                            break
                    else:
                        # If no other column found, use ID
                        generic_query = f"SELECT * FROM {table} WHERE id = ?"
            except:
                generic_query = f"SELECT * FROM {table} WHERE id = ?"
        elif query_type == "SELECT FROM ORDER BY":
            # Try to get table info to suggest a proper ORDER BY clause
            try:
                if self.db_manager:
                    table_info = self.db_manager.get_table_info(table)
                    # Find first non-ID column to use in ORDER BY clause
                    for col in table_info:
                        if col['name'].lower() != 'id':
                            generic_query = f"SELECT * FROM {table} ORDER BY {col['name']}"
                            break
                    else:
                        # If no other column found, order by ID
                        generic_query = f"SELECT * FROM {table} ORDER BY id"
            except:
                generic_query = f"SELECT * FROM {table} ORDER BY id"
        elif query_type == "Combined Query":
            # A more complex query combining several clauses
            try:
                if self.db_manager:
                    table_info = self.db_manager.get_table_info(table)
                    # Find columns to use in a more complex query
                    first_col = None
                    for col in table_info:
                        if col['name'].lower() != 'id' and 'name' in col['name'].lower():
                            first_col = col['name']
                            break
                    if first_col:
                        generic_query = f"SELECT {first_col}, COUNT(*) FROM {table} GROUP BY {first_col} ORDER BY COUNT(*) DESC LIMIT 10"
                    else:
                        generic_query = f"SELECT * FROM {table} LIMIT 10"
            except:
                generic_query = f"SELECT * FROM {table} LIMIT 10"

        query_button.clicked.connect(lambda: self.execute_specific_query(generic_query))

        # Add the button to the layout
        self.select_queries_scroll_layout.addWidget(query_button)
        self.select_queries_scroll_layout.addStretch()  # Add stretch to push content to top

    def clear_layout(self, layout):
        """Recursively clear all widgets from a layout"""
        while layout.count():
            child = layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
            elif child.layout():
                self.clear_layout(child.layout())

    def create_join_queries_tab(self):
        """Create a tab for JOIN queries"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Instructions label
        instructions = QLabel(self.get_translation("Select and execute JOIN queries between tables:"))
        instructions.setWordWrap(True)
        layout.addWidget(instructions)

        # Controls layout at the top (only selection controls, no execute button)
        controls_layout = QHBoxLayout()

        # First table selection for JOIN
        table1_label = QLabel("First table" + ":")
        controls_layout.addWidget(table1_label)

        self.join_table1_combo = QComboBox()
        self.join_table1_combo.addItems([
            "Train", "Ticket", "Service", "Schedule", "Platform", "Passenger", "Employee"
        ])
        self.join_table1_combo.currentTextChanged.connect(self.update_join_queries_display)
        controls_layout.addWidget(self.join_table1_combo)

        # Second table selection for JOIN
        table2_label = QLabel("Second table" + ":")
        controls_layout.addWidget(table2_label)

        self.join_table2_combo = QComboBox()
        self.join_table2_combo.addItems([
            "Train", "Ticket", "Service", "Schedule", "Platform", "Passenger", "Employee"
        ])
        self.join_table2_combo.currentTextChanged.connect(self.update_join_queries_display)
        controls_layout.addWidget(self.join_table2_combo)

        # JOIN type selection
        join_type_label = QLabel("JOIN type" + ":")
        controls_layout.addWidget(join_type_label)

        self.join_type_combo = QComboBox()
        self.join_type_combo.addItems([
            "INNER JOIN",
            "LEFT JOIN",
            "RIGHT JOIN",
            "FULL OUTER JOIN",
            "CROSS JOIN"
        ])
        self.join_type_combo.currentTextChanged.connect(self.update_join_queries_display)
        controls_layout.addWidget(self.join_type_combo)

        # Add stretch to push controls to the left
        controls_layout.addStretch()

        layout.addLayout(controls_layout)

        # Create scroll area for query selection
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # Store reference to the scroll layout to update dynamically
        self.join_queries_scroll_layout = scroll_layout

        # Initial population of queries
        self.update_join_queries_display()

        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        # Action buttons at the bottom
        action_layout = QHBoxLayout()

        # Save query button only (execute button removed per user request)
        self.save_join_query_btn = QPushButton(self.get_translation("save_query"))
        self.save_join_query_btn.setObjectName("save")  # Green on hover for save
        self.save_join_query_btn.clicked.connect(self.save_join_query)
        action_layout.addWidget(self.save_join_query_btn)

        # Add stretch to push buttons to the left
        action_layout.addStretch()

        layout.addLayout(action_layout)

        return widget

    def update_join_queries_display(self):
        """Update the displayed JOIN queries based on selected tables and join type"""
        # Clear existing widgets from the layout
        while self.join_queries_scroll_layout.count():
            child = self.join_queries_scroll_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
            elif child.layout():
                self.clear_layout(child.layout())

        # Get selected values
        table1 = self.join_table1_combo.currentText()
        table2 = self.join_table2_combo.currentText()
        join_type = self.join_type_combo.currentText()

        if not table1 or not table2:
            # If no tables are selected, show a message
            no_selection_label = QLabel("Please select two tables to see JOIN options")
            self.join_queries_scroll_layout.addWidget(no_selection_label)
            self.join_queries_scroll_layout.addStretch()
            return

        # Dynamically create JOIN query button based on selection
        query_button = QPushButton(f"{table1} {join_type} {table2}")
        query_button.setObjectName("ok")  # Green for execute

        # Generate a generic JOIN query since we can't rely on pre-defined queries for arbitrary tables
        generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.id = {table2}.{table1}_id"

        # Try to get better column matching information if available
        if self.db_manager:
            try:
                # Get table info for both tables
                table1_info = self.db_manager.get_table_info(table1)
                table2_info = self.db_manager.get_table_info(table2)

                # Look for potential join columns (foreign keys)
                table1_columns = [col['name'] for col in table1_info]
                table2_columns = [col['name'] for col in table2_info]

                # Simple heuristic to find potential join columns
                # Look for columns in table2 that might reference table1
                potential_join_cols = []
                for col2 in table2_columns:
                    # If column name contains table1 name followed by _id, or just id
                    if col2 == f"{table1}_id" or col2 == "id":
                        potential_join_cols.append(col2)

                # Also look for matching column names
                matching_cols = list(set(table1_columns) & set(table2_columns))

                # Construct query based on found columns
                if potential_join_cols:
                    join_col = potential_join_cols[0]
                    generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table2}.{join_col} = {table1}.id"
                elif matching_cols:
                    join_col = matching_cols[0]
                    generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.{join_col} = {table2}.{join_col}"
                else:
                    # If no obvious join column found, suggest a cross join or basic join on ID
                    if join_type == "CROSS JOIN":
                        generic_query = f"SELECT * FROM {table1} CROSS JOIN {table2}"
                    else:
                        generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.id = {table2}.id"  # Default assumption

            except Exception:
                # If we can't get table info, use the default query
                generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.id = {table2}.id"

        query_button.clicked.connect(lambda: self.execute_specific_query(generic_query))

        # Add the button to the layout
        self.join_queries_scroll_layout.addWidget(query_button)
        self.join_queries_scroll_layout.addStretch()  # Add stretch to push content to top

    def execute_join_query(self):
        """Execute a JOIN query based on selected tables and join type"""
        table1 = self.join_table1_combo.currentText()
        table2 = self.join_table2_combo.currentText()
        join_type = self.join_type_combo.currentText()

        if not table1 or not table2:
            QMessageBox.warning(self, "Warning", "Please select both tables to execute a JOIN query.")
            return

        # Generate a generic JOIN query
        generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.id = {table2}.{table1}_id"

        # Try to get better column matching information if available
        if self.db_manager:
            try:
                # Get table info for both tables
                table1_info = self.db_manager.get_table_info(table1)
                table2_info = self.db_manager.get_table_info(table2)

                # Look for potential join columns (foreign keys)
                table1_columns = [col['name'] for col in table1_info]
                table2_columns = [col['name'] for col in table2_info]

                # Simple heuristic to find potential join columns
                potential_join_cols = []
                for col2 in table2_columns:
                    # If column name contains table1 name followed by _id, or just id
                    if col2 == f"{table1}_id" or col2 == "id":
                        potential_join_cols.append(col2)

                # Also look for matching column names
                matching_cols = list(set(table1_columns) & set(table2_columns))

                # Construct query based on found columns
                if potential_join_cols:
                    join_col = potential_join_cols[0]
                    generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table2}.{join_col} = {table1}.id"
                elif matching_cols:
                    join_col = matching_cols[0]
                    generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.{join_col} = {table2}.{join_col}"
                else:
                    # If no obvious join column found, suggest a cross join or basic join on ID
                    if join_type == "CROSS JOIN":
                        generic_query = f"SELECT * FROM {table1} CROSS JOIN {table2}"
                    else:
                        generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.id = {table2}.id"  # Default assumption

            except Exception:
                # If we can't get table info, use the default query
                generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.id = {table2}.id"

        self.execute_query_and_display(generic_query)

    def save_join_query(self):
        """Save the currently selected JOIN query"""
        table1 = self.join_table1_combo.currentText()
        table2 = self.join_table2_combo.currentText()
        join_type = self.join_type_combo.currentText()

        if not table1 or not table2:
            QMessageBox.warning(self, "Warning", "Please select both tables first.")
            return

        # Generate the query to save
        generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.id = {table2}.{table1}_id"

        # Try to get better column matching information if available
        if self.db_manager:
            try:
                # Get table info for both tables
                table1_info = self.db_manager.get_table_info(table1)
                table2_info = self.db_manager.get_table_info(table2)

                # Look for potential join columns (foreign keys)
                table1_columns = [col['name'] for col in table1_info]
                table2_columns = [col['name'] for col in table2_info]

                # Simple heuristic to find potential join columns
                potential_join_cols = []
                for col2 in table2_columns:
                    if col2 == f"{table1}_id" or col2 == "id":
                        potential_join_cols.append(col2)

                # Also look for matching column names
                matching_cols = list(set(table1_columns) & set(table2_columns))

                if potential_join_cols:
                    join_col = potential_join_cols[0]
                    generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table2}.{join_col} = {table1}.id"
                elif matching_cols:
                    join_col = matching_cols[0]
                    generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.{join_col} = {table2}.{join_col}"
                else:
                    if join_type == "CROSS JOIN":
                        generic_query = f"SELECT * FROM {table1} CROSS JOIN {table2}"
                    else:
                        generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.id = {table2}.id"

            except Exception:
                generic_query = f"SELECT * FROM {table1} {join_type} {table2} ON {table1}.id = {table2}.id"

        # Ask user for a name/description for the query
        query_name, ok = QInputDialog.getText(
            self,
            "Save Query",
            "Enter a name for this query:",
            text=f"{table1} {join_type} {table2}"
        )

        if ok and query_name:
            # Save the query to the query history manager
            self.query_history_manager.add_query(generic_query, query_name)  # query first, then description
            QMessageBox.information(self, "Success", f"Query '{query_name}' saved successfully!")

    def create_aggregate_queries_tab(self):
        """Create a tab for aggregate queries, subqueries, and set operations"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Instructions label
        instructions = QLabel(self.get_translation("Select and execute aggregate, subquery, and set operation queries:"))
        instructions.setWordWrap(True)
        layout.addWidget(instructions)

        # Create scroll area for query selection
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # Aggregate and subquery buttons
        agg_label = QLabel("Aggregate and Subquery Operations:")
        scroll_layout.addWidget(agg_label)

        agg_layout = QHBoxLayout()
        train_agg_btn = QPushButton("Train: AVG, COUNT")
        train_agg_btn.setObjectName("purple")  # Purple for Train aggregate
        train_agg_btn.clicked.connect(lambda: self.execute_specific_query(AggregateQueries.AGGREGATE_TRAIN))
        ticket_agg_btn = QPushButton("Ticket: SUM, MIN, MAX")
        ticket_agg_btn.setObjectName("red")  # Red for Ticket aggregate
        ticket_agg_btn.clicked.connect(lambda: self.execute_specific_query(AggregateQueries.AGGREGATE_TICKET))
        service_agg_btn = QPushButton("Service: GROUP BY, HAVING")
        service_agg_btn.setObjectName("orange")  # Orange for Service aggregate
        service_agg_btn.clicked.connect(lambda: self.execute_specific_query(AggregateQueries.AGGREGATE_SERVICE))
        agg_layout.addWidget(train_agg_btn)
        agg_layout.addWidget(ticket_agg_btn)
        agg_layout.addWidget(service_agg_btn)
        scroll_layout.addLayout(agg_layout)

        subquery_label = QLabel("Subquery Operations:")
        scroll_layout.addWidget(subquery_label)

        subquery_btn = QPushButton("Schedule: Subqueries in WHERE")
        subquery_btn.clicked.connect(lambda: self.execute_specific_query(AggregateQueries.SUBQUERY_SCHEDULE))
        scroll_layout.addWidget(subquery_btn)

        set_op_label = QLabel("Set Operations:")
        scroll_layout.addWidget(set_op_label)

        set_layout = QHBoxLayout()
        union_btn = QPushButton("UNION: Platform & Train")
        union_btn.clicked.connect(lambda: self.execute_specific_query(AggregateQueries.UNION_PLATFORM_TRAIN))
        intersect_btn = QPushButton("INTERSECT: Passenger & Ticket")
        intersect_btn.clicked.connect(lambda: self.execute_specific_query(AggregateQueries.INTERSECT_PASSENGER_TICKET))
        except_btn = QPushButton("EXCEPT: Employee & Service")
        except_btn.clicked.connect(lambda: self.execute_specific_query(AggregateQueries.EXCEPT_EMPLOYEE_SERVICE))
        set_layout.addWidget(union_btn)
        set_layout.addWidget(intersect_btn)
        set_layout.addWidget(except_btn)
        scroll_layout.addLayout(set_layout)

        # Add some stretch to push content to top
        scroll_layout.addStretch()

        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        return widget

    def execute_select_query(self):
        """Execute SELECT query based on selected table and query type"""
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        table = self.select_table_combo.currentText()
        query_type = self.select_query_type_combo.currentText()

        if not table:
            QMessageBox.warning(self, "Warning", "Please select a table to execute a query.")
            return

        # Generate query based on selected table and query type
        generic_query = ""

        if query_type == "SELECT FROM":
            generic_query = f"SELECT * FROM {table}"
        elif query_type == "SELECT FROM WHERE":
            # Try to get table info to suggest a proper WHERE clause
            try:
                table_info = self.db_manager.get_table_info(table)
                # Find first non-ID column to use in WHERE clause
                for col in table_info:
                    if col['name'].lower() != 'id':
                        generic_query = f"SELECT * FROM {table} WHERE {col['name']} = ?"
                        break
                else:
                    # If no other column found, use ID
                    generic_query = f"SELECT * FROM {table} WHERE id = ?"
            except:
                generic_query = f"SELECT * FROM {table} WHERE id = ?"
        elif query_type == "SELECT FROM ORDER BY":
            # Try to get table info to suggest a proper ORDER BY clause
            try:
                table_info = self.db_manager.get_table_info(table)
                # Find first non-ID column to use in ORDER BY clause
                for col in table_info:
                    if col['name'].lower() != 'id':
                        generic_query = f"SELECT * FROM {table} ORDER BY {col['name']}"
                        break
                else:
                    # If no other column found, order by ID
                    generic_query = f"SELECT * FROM {table} ORDER BY id"
            except:
                generic_query = f"SELECT * FROM {table} ORDER BY id"
        elif query_type == "Combined Query":
            # A more complex query combining several clauses
            try:
                table_info = self.db_manager.get_table_info(table)
                # Find columns to use in a more complex query
                first_col = None
                for col in table_info:
                    if col['name'].lower() != 'id' and 'name' in col['name'].lower():
                        first_col = col['name']
                        break
                if first_col:
                    generic_query = f"SELECT {first_col}, COUNT(*) FROM {table} GROUP BY {first_col} ORDER BY COUNT(*) DESC LIMIT 10"
                else:
                    generic_query = f"SELECT * FROM {table} LIMIT 10"
            except:
                generic_query = f"SELECT * FROM {table} LIMIT 10"

        if generic_query:
            self.execute_query_and_display(generic_query)

    def save_select_query(self):
        """Save the currently selected SELECT query"""
        table = self.select_table_combo.currentText()
        query_type = self.select_query_type_combo.currentText()

        if not table:
            QMessageBox.warning(self, "Warning", "Please select a table first.")
            return

        # Generate the query to save
        generic_query = ""

        if query_type == "SELECT FROM":
            generic_query = f"SELECT * FROM {table}"
        elif query_type == "SELECT FROM WHERE":
            try:
                if self.db_manager:
                    table_info = self.db_manager.get_table_info(table)
                    for col in table_info:
                        if col['name'].lower() != 'id':
                            generic_query = f"SELECT * FROM {table} WHERE {col['name']} = ?"
                            break
                    else:
                        generic_query = f"SELECT * FROM {table} WHERE id = ?"
            except:
                generic_query = f"SELECT * FROM {table} WHERE id = ?"
        elif query_type == "SELECT FROM ORDER BY":
            try:
                if self.db_manager:
                    table_info = self.db_manager.get_table_info(table)
                    for col in table_info:
                        if col['name'].lower() != 'id':
                            generic_query = f"SELECT * FROM {table} ORDER BY {col['name']}"
                            break
                    else:
                        generic_query = f"SELECT * FROM {table} ORDER BY id"
            except:
                generic_query = f"SELECT * FROM {table} ORDER BY id"
        elif query_type == "Combined Query":
            try:
                if self.db_manager:
                    table_info = self.db_manager.get_table_info(table)
                    first_col = None
                    for col in table_info:
                        if col['name'].lower() != 'id' and 'name' in col['name'].lower():
                            first_col = col['name']
                            break
                    if first_col:
                        generic_query = f"SELECT {first_col}, COUNT(*) FROM {table} GROUP BY {first_col} ORDER BY COUNT(*) DESC LIMIT 10"
                    else:
                        generic_query = f"SELECT * FROM {table} LIMIT 10"
            except:
                generic_query = f"SELECT * FROM {table} LIMIT 10"

        # Ask user for a name/description for the query
        query_name, ok = QInputDialog.getText(
            self,
            "Save Query",
            "Enter a name for this query:",
            text=f"{table} {query_type}"
        )

        if ok and query_name:
            # Save the query to the query history manager
            self.query_history_manager.add_query(generic_query, query_name)  # query first, then description
            QMessageBox.information(self, "Success", f"Query '{query_name}' saved successfully!")

    def update_table_combos(self):
        """Update table selection combos with actual database tables"""
        if not self.db_manager:
            return

        try:
            tables = self.db_manager.get_tables()

            # Update SELECT queries tab combo (unified tab)
            if hasattr(self, 'select_table_combo'):
                current_text = self.select_table_combo.currentText()
                self.select_table_combo.clear()
                self.select_table_combo.addItems(tables)
                # Try to preserve previous selection if it still exists
                if current_text in tables:
                    self.select_table_combo.setCurrentText(current_text)
                elif tables:  # If no previous selection exists but tables exist, select first table
                    self.select_table_combo.setCurrentText(tables[0])  # This will trigger the signal connection


            # Update JOIN queries tab combos (unified tab)
            if hasattr(self, 'join_table1_combo'):
                current_text1 = self.join_table1_combo.currentText()
                self.join_table1_combo.clear()
                self.join_table1_combo.addItems(tables)
                if current_text1 in tables:
                    self.join_table1_combo.setCurrentText(current_text1)
                elif tables:  # If no previous selection exists but tables exist, select first table
                    self.join_table1_combo.setCurrentText(tables[0])  # This will trigger the signal connection

            if hasattr(self, 'join_table2_combo'):
                current_text2 = self.join_table2_combo.currentText()
                self.join_table2_combo.clear()
                self.join_table2_combo.addItems(tables)
                if current_text2 in tables:
                    self.join_table2_combo.setCurrentText(current_text2)
                elif tables:  # If no previous selection exists but tables exist, select first table
                    self.join_table2_combo.setCurrentText(tables[0])  # This will trigger the signal connection

            # Update Aggregate queries tab combo (unified tab)
            if hasattr(self, 'agg_table_combo'):
                current_text_agg = self.agg_table_combo.currentText()
                self.agg_table_combo.clear()
                self.agg_table_combo.addItems(tables)
                if current_text_agg in tables:
                    self.agg_table_combo.setCurrentText(current_text_agg)
                elif tables:  # If no previous selection exists but tables exist, select first table
                    self.agg_table_combo.setCurrentText(tables[0])  # This will trigger the signal connection

            # For additional safety, make sure the column dropdowns are updated if a table is selected
            # This ensures that even if the signal didn't fire properly, the columns are updated
            if hasattr(self, 'select_table_combo') and tables:
                selected_table = self.select_table_combo.currentText()
                if selected_table in tables:
                    self.update_table_columns(selected_table)

        except Exception as e:
            print(f"Error updating table combos: {e}")

    def execute_specific_query(self, query):
        """Execute a specific query passed as argument"""
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        self.execute_query_and_display(query)

    def execute_query_and_display(self, query):
        """Execute the given query and display results"""
        try:
            # Switch to the Query Editor tab to show results
            self.tab_widget.setCurrentIndex(1)  # Index 1 is the query editor tab

            # Display the query in the input field
            self.query_input.setPlainText(query)

            result = self.db_manager.execute_query(query)

            if result is not None:  # SELECT query
                if result:
                    # Get column names
                    columns = [desc[0] for desc in self.db_manager.cursor.description]

                    # Set up the results table
                    self.query_results.setRowCount(len(result))
                    self.query_results.setColumnCount(len(columns))
                    self.query_results.setHorizontalHeaderLabels(columns)

                    # Populate the table with color coding for data types
                    for i, row in enumerate(result):
                        for j, value in enumerate(row):
                            item = QTableWidgetItem(str(value))

                            # Determine data type and apply appropriate styling
                            value_str = str(value)
                            if value_str.isdigit() or (value_str.startswith('-') and value_str[1:].isdigit()) or self.is_numeric(value_str):
                                # Numeric data - greenish
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_NUMERIC)
                                item.setBackground(QColor(COLOR_NUMERIC_BG))
                                item.setForeground(QColor(COLOR_NUMERIC_FG))
                            elif self.is_date_or_time(value_str):
                                # Date/time data - orange
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_DATE)
                                item.setBackground(QColor(COLOR_DATE_BG))
                                item.setForeground(QColor(COLOR_DATE_FG))
                            elif value_str.lower() in ['true', 'false', '1', '0', 'yes', 'no']:
                                # Boolean data - teal
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_BOOLEAN)
                                item.setBackground(QColor(COLOR_BOOLEAN_BG))
                                item.setForeground(QColor(COLOR_BOOLEAN_FG))
                            else:
                                # Text data - blueish
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_TEXT)
                                item.setBackground(QColor(COLOR_TEXT_BG))
                                item.setForeground(QColor(COLOR_TEXT_FG))

                            self.query_results.setItem(i, j, item)

                    # Resize columns to fit content
                    self.query_results.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
                else:
                    # No results but query executed successfully
                    self.query_results.setRowCount(0)
                    self.query_results.setColumnCount(0)

                self.statusBar().showMessage(self.get_translation("query_executed_successfully_rows").format(count=len(result) if result else 0))
            else:  # Non-SELECT query
                self.query_results.setRowCount(0)
                self.query_results.setColumnCount(0)
                self.statusBar().showMessage(self.get_translation("query_executed_successfully"))

            # Add the query to history with a descriptive name
            query_description = f"Executed query: {query.strip()[:50]}{'...' if len(query.strip()) > 50 else ''}"
            self.query_history_manager.add_query(query, query_description)
            self.refresh_query_history()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error executing query: {str(e)}")

    def undo_action_triggered(self):
        """Handle undo action triggered"""
        if self.backup_manager:
            if self.backup_manager.undo():
                self.statusBar().showMessage(self.get_translation("Undo completed"))
                self.update_undo_redo_states()
            else:
                self.statusBar().showMessage(self.get_translation("Nothing to undo"))
        else:
            self.statusBar().showMessage(self.get_translation("Backup manager not initialized"))

    def redo_action_triggered(self):
        """Handle redo action triggered"""
        if self.backup_manager:
            if self.backup_manager.redo():
                self.statusBar().showMessage(self.get_translation("Redo completed"))
                self.update_undo_redo_states()
            else:
                self.statusBar().showMessage(self.get_translation("Nothing to redo"))
        else:
            self.statusBar().showMessage(self.get_translation("Backup manager not initialized"))

    def update_undo_redo_states(self):
        """Update the enabled state of undo/redo actions"""
        if self.backup_manager:
            can_undo = self.backup_manager.can_undo()
            can_redo = self.backup_manager.can_redo()

            self.undo_action.setEnabled(can_undo)
            self.redo_action.setEnabled(can_redo)

            # Update tooltips with action descriptions
            undo_desc = self.backup_manager.get_undo_description()
            redo_desc = self.backup_manager.get_redo_description()

            self.undo_action.setText(f"Undo: {undo_desc}" if undo_desc else "Undo")
            self.redo_action.setText(f"Redo: {redo_desc}" if redo_desc else "Redo")

    def connect_to_database(self):
        """Show connection dialog to connect to PostgreSQL database"""
        from ui.connection_dialog import ConnectionDialog

        dialog = ConnectionDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            connection_info = dialog.get_connection_info()
            db_type = connection_info['db_type']
            connection_params = connection_info['connection_params']

            if self.db_manager:
                self.db_manager.close()

            self.db_manager = DatabaseManager(
                db_type=db_type,
                connection_params=connection_params
            )

            if self.db_manager.connect():
                db_identifier = connection_params.get('database', connection_params.get('path', 'Unknown'))
                if db_type == 'postgresql':
                    db_identifier = f"PostgreSQL_{connection_params.get('database', 'db')}@{connection_params.get('host', 'localhost')}"

                self.statusBar().showMessage(self.get_translation("connected_to_database").format(database=db_identifier))
                self.backup_manager = BackupManager(self.db_manager)

                # Set the database for query history manager to save/load queries specific to this database
                db_path = f"{db_type}_{db_identifier}"
                self.query_history_manager.set_database(db_path)

                self.refresh_tables()
                self.update_table_combos()  # Update table combos with actual database tables
                self.refresh_saved_queries()  # Refresh saved queries list after connection
                return True
            else:
                QMessageBox.critical(self, "Connection Error", self.get_translation("Failed to connect to the database."))
                return False

    def connect_to_postgresql(self, connection_params):
        """Connect directly to PostgreSQL database with provided parameters"""
        if self.db_manager:
            self.db_manager.close()

        self.db_manager = DatabaseManager(
            db_type='postgresql',
            connection_params=connection_params
        )

        if self.db_manager.connect():
            db_identifier = f"PostgreSQL_{connection_params.get('database', 'db')}@{connection_params.get('host', 'localhost')}_{connection_params.get('port', 5432)}"

            self.statusBar().showMessage(self.get_translation("connected_to_database").format(database=db_identifier))
            self.backup_manager = BackupManager(self.db_manager)

            # Set the database for query history manager to save/load queries specific to this database
            db_path = f"postgresql_{db_identifier}"
            self.query_history_manager.set_database(db_path)

            self.refresh_tables()
            self.update_table_combos()  # Update table combos with actual database tables
            self.refresh_saved_queries()  # Refresh saved queries list after connection
            return True
        else:
            QMessageBox.critical(self, "Connection Error", self.get_translation("Failed to connect to the database."))
            return False


    def refresh_tables(self):
        if not self.db_manager:
            return

        tables = self.db_manager.get_tables()

        self.tables_list.setRowCount(len(tables))
        for i, table in enumerate(tables):
            self.tables_list.setItem(i, 0, QTableWidgetItem(table))

        # Clear columns list
        self.columns_list.setRowCount(0)

        # Refresh query history
        self.refresh_query_history()

    def table_selected(self, row, column):
        if not self.db_manager:
            return

        table_name = self.tables_list.item(row, 0).text()

        # Initialize pagination for the selected table
        self.update_pagination_info(table_name)

        # Load first page of data
        self.load_table_data(table_name, limit=int(self.rows_per_page_combo.currentText()), offset=0)

        # Load table columns
        self.load_table_columns(table_name)

    def load_table_data(self, table_name, limit=1000, offset=0):
        """
        Load table data with performance optimization and pagination support.

        Args:
            table_name (str): Name of the table to load data from
            limit (int): Maximum number of rows to load (default 1000)
            offset (int): Number of rows to skip (for pagination)
        """
        try:
            # First, check if the table has too many rows to load at once
            count_query = f"SELECT COUNT(*) FROM {table_name}"
            count_result = self.db_manager.execute_query(count_query)
            total_rows = count_result[0][0] if count_result and count_result[0] else 0

            if total_rows > limit:
                # Show warning if table is large
                reply = QMessageBox.question(
                    self,
                    "Large Table Warning",
                    f"The table '{table_name}' has {total_rows} rows. "
                    f"Loading all rows may impact performance. "
                    f"Do you want to load the first {limit} rows?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )

                if reply == QMessageBox.StandardButton.No:
                    return

            # Load data with limit
            query = f"SELECT * FROM {table_name} LIMIT {limit} OFFSET {offset}"
            data = self.db_manager.execute_query(query)

            if not data:
                self.table_view.setRowCount(0)
                self.table_view.setColumnCount(0)
                return

            # Get column names
            columns = [desc[0] for desc in self.db_manager.cursor.description]

            # Temporarily disable updates for performance
            self.table_view.setUpdatesEnabled(False)
            self.table_view.setSortingEnabled(False)  # Disable sorting during population

            # Set up the table
            self.table_view.setRowCount(len(data))
            self.table_view.setColumnCount(len(columns))
            self.table_view.setHorizontalHeaderLabels(columns)

            # Populate the table in chunks for better performance
            chunk_size = 100  # Process in chunks of 100 rows
            for i, row in enumerate(data):
                # Process in chunks to allow UI updates
                if i % chunk_size == 0:
                    QApplication.processEvents()  # Allow UI to update

                for j, value in enumerate(row):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    # Apply highlighting based on data type
                    self.apply_data_type_highlighting(item, str(value) if value is not None else "")
                    self.table_view.setItem(i, j, item)

            # Re-enable updates
            self.table_view.setUpdatesEnabled(True)
            self.table_view.setSortingEnabled(True)  # Re-enable sorting after population

            # Resize columns to fit content (more efficiently)
            header = self.table_view.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

            # Show status message with row count
            self.statusBar().showMessage(f"Loaded {len(data)} rows from {table_name}. Total rows in table: {total_rows}")

        except Exception as e:
            error_msg = f"Failed to load table data: {str(e)}"
            self.logger.error(error_msg)
            QMessageBox.critical(self, "Error", error_msg)

    def load_table_columns(self, table_name):
        try:
            columns = self.db_manager.get_table_info(table_name)

            self.columns_list.setRowCount(len(columns))
            for i, col in enumerate(columns):
                self.columns_list.setItem(i, 0, QTableWidgetItem(col['name']))
                self.columns_list.setItem(i, 1, QTableWidgetItem(col['type']))
                self.columns_list.setItem(i, 2, QTableWidgetItem("Yes" if col['not_null'] else "No"))
                self.columns_list.setItem(i, 3, QTableWidgetItem(str(col['default']) if col['default'] else ""))
                self.columns_list.setItem(i, 4, QTableWidgetItem("Yes" if col['primary_key'] else "No"))

            # Also populate the filter column combo
            self.filter_column_combo.clear()
            for col in columns:
                self.filter_column_combo.addItem(col['name'])
        except Exception as e:
            print(f"Error loading table columns: {e}")
            QMessageBox.critical(self, "Error", f"Error loading table columns: {str(e)}")

    def apply_table_filter(self):
        """Apply filter to the current table and show results in Query Editor"""
        # Get the currently selected table
        current_row = self.tables_list.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Warning", "Please select a table first.")
            return

        table_name = self.tables_list.item(current_row, 0).text()
        if not table_name:
            QMessageBox.warning(self, "Warning", "Please select a table first.")
            return

        # Get selected column and filter value
        column_name = self.filter_column_combo.currentText()
        filter_value = self.filter_value_input.text().strip()

        if not column_name:
            QMessageBox.warning(self, "Warning", "Please select a column to filter by.")
            return

        if not filter_value:
            QMessageBox.warning(self, "Warning", "Please enter a filter value.")
            return

        # Create the SQL query with filter
        query = f"SELECT * FROM {table_name} WHERE {column_name} LIKE '%{filter_value}%'"

        # Switch to the Query Editor tab to show results
        self.tab_widget.setCurrentIndex(2)  # Index 2 is the query editor tab

        # Display the query in the input field
        self.query_input.setPlainText(query)

        try:
            # Execute the query
            result = self.db_manager.execute_query(query)

            if result is not None:  # SELECT query
                if result:
                    # Get column names
                    columns = [desc[0] for desc in self.db_manager.cursor.description]

                    # Set up the results table
                    self.query_results.setRowCount(len(result))
                    self.query_results.setColumnCount(len(columns))
                    self.query_results.setHorizontalHeaderLabels(columns)

                    # Populate the table with color coding for data types
                    for i, row in enumerate(result):
                        for j, value in enumerate(row):
                            item = QTableWidgetItem(str(value))

                            # Determine data type and apply appropriate styling
                            value_str = str(value)
                            if value_str.isdigit() or (value_str.startswith('-') and value_str[1:].isdigit()) or self.is_numeric(value_str):
                                # Numeric data - greenish
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_NUMERIC)
                                item.setBackground(QColor(COLOR_NUMERIC_BG))
                                item.setForeground(QColor(COLOR_NUMERIC_FG))
                            elif self.is_date_or_time(value_str):
                                # Date/time data - orange
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_DATE)
                                item.setBackground(QColor(COLOR_DATE_BG))
                                item.setForeground(QColor(COLOR_DATE_FG))
                            elif value_str.lower() in ['true', 'false', '1', '0', 'yes', 'no']:
                                # Boolean data - teal
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_BOOLEAN)
                                item.setBackground(QColor(COLOR_BOOLEAN_BG))
                                item.setForeground(QColor(COLOR_BOOLEAN_FG))
                            else:
                                # Text data - blueish
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_TEXT)
                                item.setBackground(QColor(COLOR_TEXT_BG))
                                item.setForeground(QColor(COLOR_TEXT_FG))

                            self.query_results.setItem(i, j, item)

                    # Resize columns to fit content
                    self.query_results.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
                else:
                    # No results but query executed successfully
                    self.query_results.setRowCount(0)
                    self.query_results.setColumnCount(0)

                self.statusBar().showMessage(f"Filter applied successfully. Found {len(result) if result else 0} rows.")
            else:  # Non-SELECT query (shouldn't happen with filter query)
                self.query_results.setRowCount(0)
                self.query_results.setColumnCount(0)
                self.statusBar().showMessage("Filter query executed successfully.")

            # Add the query to history
            query_description = f"Filter query on table '{table_name}' by column '{column_name}' with value '{filter_value}'"
            self.query_history_manager.add_query(query, query_description)
            self.refresh_query_history()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error executing filter query: {str(e)}")

    def clear_table_filter(self):
        """Clear the current filter"""
        self.filter_column_combo.clear()
        self.filter_value_input.clear()

        # Get the currently selected table to reload columns
        current_row = self.tables_list.currentRow()
        if current_row >= 0:
            table_name = self.tables_list.item(current_row, 0).text()
            if table_name:
                self.load_table_columns(table_name)

        self.statusBar().showMessage("Filter cleared.")

    def update_pagination_info(self, table_name, limit=1000):
        """Update pagination controls based on table size"""
        try:
            # Get total row count
            count_query = f"SELECT COUNT(*) FROM {table_name}"
            count_result = self.db_manager.execute_query(count_query)
            total_rows = count_result[0][0] if count_result and count_result[0] else 0

            # Calculate total pages
            current_limit = int(self.rows_per_page_combo.currentText())
            total_pages = max(1, (total_rows + current_limit - 1) // current_limit)

            # Update pagination controls
            self.page_spinbox.setMaximum(total_pages)
            self.total_pages_label.setText(f"/ {total_pages}")

            # Update status bar
            self.statusBar().showMessage(f"Table '{table_name}' has {total_rows} rows. Showing {current_limit} per page. Total pages: {total_pages}")

        except Exception as e:
            self.logger.error(f"Error updating pagination info: {str(e)}")

    def go_to_first_page(self):
        """Navigate to the first page"""
        if self.page_spinbox.value() > 1:
            self.page_spinbox.setValue(1)
            self.load_current_page()

    def go_to_prev_page(self):
        """Navigate to the previous page"""
        if self.page_spinbox.value() > 1:
            self.page_spinbox.setValue(self.page_spinbox.value() - 1)
            self.load_current_page()

    def go_to_next_page(self):
        """Navigate to the next page"""
        max_page = self.page_spinbox.maximum()
        if self.page_spinbox.value() < max_page:
            self.page_spinbox.setValue(self.page_spinbox.value() + 1)
            self.load_current_page()

    def go_to_last_page(self):
        """Navigate to the last page"""
        max_page = self.page_spinbox.maximum()
        if self.page_spinbox.value() < max_page:
            self.page_spinbox.setValue(max_page)
            self.load_current_page()

    def load_current_page(self):
        """Load the currently selected page"""
        current_row = self.tables_list.currentRow()
        if current_row >= 0:
            table_name = self.tables_list.item(current_row, 0).text()
            if table_name:
                current_page = self.page_spinbox.value()
                rows_per_page = int(self.rows_per_page_combo.currentText())
                offset = (current_page - 1) * rows_per_page

                self.load_table_data(table_name, limit=rows_per_page, offset=offset)

                # Update status bar with current page info
                self.statusBar().showMessage(f"Showing page {current_page} of {self.page_spinbox.maximum()} for table '{table_name}'")

    def refresh_current_table(self):
        # Get the currently selected table
        selected_items = self.tables_list.selectedItems()
        if selected_items:
            table_name = selected_items[0].text()

            # Update pagination info and refresh current page
            self.update_pagination_info(table_name)
            self.load_current_page()

    def add_new_table(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        table_name, ok = QInputDialog.getText(self, self.get_translation("Add New Table"), self.get_translation("Enter table name:"))
        if ok and table_name:
            try:
                # Create undo action for table creation
                def undo_func():
                    # Drop the created table to undo its creation
                    drop_query = f"DROP TABLE IF EXISTS {table_name}"
                    self.db_manager.execute_query(drop_query)
                    self.refresh_tables()
                    self.statusBar().showMessage(f"Undo: Table '{table_name}' creation reverted")

                def redo_func():
                    # Recreate the table to redo the action
                    if self.db_manager.db_type == 'postgresql':
                        query = f"CREATE TABLE {table_name} (id SERIAL PRIMARY KEY)"
                    else:
                        query = f"CREATE TABLE {table_name} (id INTEGER PRIMARY KEY AUTOINCREMENT)"
                    self.db_manager.execute_query(query)
                    self.refresh_tables()
                    self.statusBar().showMessage(self.get_translation("table_created_successfully").format(table=table_name))

                # Execute the table creation with database-specific syntax
                if self.db_manager.db_type == 'postgresql':
                    query = f"CREATE TABLE {table_name} (id SERIAL PRIMARY KEY)"
                else:
                    query = f"CREATE TABLE {table_name} (id INTEGER PRIMARY KEY AUTOINCREMENT)"
                self.db_manager.execute_query(query)
                self.refresh_tables()

                # Add to action history for undo/redo
                if self.backup_manager:
                    self.backup_manager.create_db_operation_action(
                        operation_name="create_table",
                        undo_func=undo_func,
                        redo_func=redo_func,
                        description=f"Create table '{table_name}'"
                    )
                    self.update_undo_redo_states()

                self.statusBar().showMessage(self.get_translation("table_created_successfully").format(table=table_name))
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to create table: {str(e)}")

    def delete_table(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        selected_items = self.tables_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", self.get_translation("Please select a table to delete."))
            return

        table_name = selected_items[0].text()

        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete table '{table_name}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Get the table's structure and data for potential undo
                # Use PostgreSQL-specific query instead of PRAGMA
                table_info = self.db_manager.get_table_info(table_name)

                # Get the table's data for potential undo
                data_query = f"SELECT * FROM {table_name}"
                data_result = self.db_manager.execute_query(data_query)

                # Create undo action for table deletion
                def undo_func():
                    # Recreate the table structure if it doesn't exist
                    if not self.db_manager.table_exists(table_name) and table_info:
                        # Build the CREATE TABLE query from the saved table info
                        columns_def = []
                        for col_info in table_info:
                            col_name = col_info[1]  # column name
                            col_type = col_info[2]  # column type
                            not_null = "NOT NULL" if col_info[3] else ""
                            primary_key = "PRIMARY KEY" if col_info[5] else ""
                            default_val = f"DEFAULT '{col_info[4]}'" if col_info[4] else ""

                            col_def = f"{col_name} {col_type}"
                            if not_null:
                                col_def += f" {not_null}"
                            if primary_key:
                                col_def += f" {primary_key}"
                            if default_val and "DEFAULT" in default_val:
                                col_def += f" {default_val}"

                            columns_def.append(col_def)

                        create_query = f"CREATE TABLE {table_name} ({', '.join(columns_def)})"
                        self.db_manager.execute_query(create_query)

                    # Insert the saved data back
                    if data_result and table_info:
                        column_names = [info[1] for info in table_info]
                        placeholders = ", ".join(["?" for _ in column_names])
                        insert_query = f"INSERT INTO {table_name} ({', '.join(column_names)}) VALUES ({placeholders})"
                        for row in data_result:
                            self.db_manager.execute_query(insert_query, row)

                    self.refresh_tables()
                    self.statusBar().showMessage(f"Undo: Table '{table_name}' restored")

                def redo_func():
                    # Drop the table again to redo the action
                    query = f"DROP TABLE {table_name}"
                    self.db_manager.execute_query(query)
                    self.refresh_tables()
                    self.statusBar().showMessage(self.get_translation("table_deleted_successfully").format(table=table_name))

                # Execute the table deletion
                query = f"DROP TABLE {table_name}"
                self.db_manager.execute_query(query)
                self.refresh_tables()

                # Add to action history for undo/redo
                if self.backup_manager:
                    self.backup_manager.create_db_operation_action(
                        operation_name="delete_table",
                        undo_func=undo_func,
                        redo_func=redo_func,
                        description=f"Delete table '{table_name}'"
                    )
                    self.update_undo_redo_states()

                self.statusBar().showMessage(self.get_translation("table_deleted_successfully").format(table=table_name))
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete table: {str(e)}")

    def add_column(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        selected_items = self.tables_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", self.get_translation("Please select a table first."))
            return

        table_name = selected_items[0].text()

        column_name, ok = QInputDialog.getText(self, self.get_translation("Add Column"), self.get_translation("Enter column name:"))
        if not ok or not column_name:
            return

        data_type, ok = QInputDialog.getItem(
            self,
            self.get_translation("Column Data Type"),
            self.get_translation("Select data type:"),
            ["TEXT", "INTEGER", "REAL", "BLOB", "NUMERIC", "BOOLEAN", "DATE", "TIME", "TIMESTAMP"],
            0, False
        )
        if not ok:
            return

        try:
            query = f"ALTER TABLE {table_name} ADD COLUMN {column_name} {data_type}"
            self.db_manager.execute_query(query)
            self.load_table_columns(table_name)
            self.statusBar().showMessage(self.get_translation("column_added_to_table").format(column=column_name, table=table_name))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add column: {str(e)}")

    def delete_column(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        selected_items = self.tables_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", self.get_translation("Please select a table first."))
            return

        table_name = selected_items[0].text()

        # Get columns for the table
        try:
            columns = self.db_manager.get_table_info(table_name)
            column_names = [col['name'] for col in columns]

            if not column_names:
                QMessageBox.warning(self, "Warning", f"Table '{table_name}' has no columns.")
                return

            column_name, ok = QInputDialog.getItem(
                self,
                self.get_translation("Delete Column"),
                self.get_translation("Select column to delete:"),
                column_names,
                0, False
            )

            if ok and column_name:
                reply = QMessageBox.question(
                    self,
                    "Confirm Delete",
                    f"Are you sure you want to delete column '{column_name}'?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )

                if reply == QMessageBox.StandardButton.Yes:
                    # For PostgreSQL, we can directly drop columns
                    try:
                        alter_query = f"ALTER TABLE {table_name} DROP COLUMN {column_name};"
                        self.db_manager.execute_query(alter_query)
                        QMessageBox.information(self, "Success", f"Column '{column_name}' deleted successfully.")
                        self.refresh_tables()
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to delete column: {str(e)}")
                    remaining_columns = [col for col in columns if col['name'] != column_name]
                    if not remaining_columns:
                        QMessageBox.warning(self, "Warning", self.get_translation("You cannot delete the last column."))
                        return

                    # Get data from the table
                    query = f"SELECT * FROM {table_name}"
                    data = self.db_manager.execute_query(query)

                    # Drop the old table
                    self.db_manager.execute_query(f"DROP TABLE {table_name}")

                    # Create new table without the deleted column
                    new_columns_def = ", ".join([f"{col['name']} {col['type']}" for col in remaining_columns])
                    create_query = f"CREATE TABLE {table_name} ({new_columns_def})"
                    self.db_manager.execute_query(create_query)

                    # Insert the data back
                    if data:
                        placeholders = ", ".join(["?" for _ in remaining_columns])
                        insert_query = f"INSERT INTO {table_name} VALUES ({placeholders})"
                        for row in data:
                            # Exclude the deleted column from the row
                            new_row = []
                            for i, col in enumerate(columns):
                                if col['name'] != column_name:
                                    new_row.append(row[i])
                            self.db_manager.execute_query(insert_query, tuple(new_row))

                    self.load_table_columns(table_name)
                    self.load_table_data(table_name, limit=1000, offset=0)
                    self.statusBar().showMessage(self.get_translation("column_deleted_from_table").format(column=column_name, table=table_name))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to delete column: {str(e)}")

    def edit_column(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        selected_items = self.tables_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", self.get_translation("Please select a table first."))
            return

        table_name = selected_items[0].text()

        # Get current columns
        columns_info = self.db_manager.get_table_info(table_name)
        if not columns_info:
            QMessageBox.warning(self, "Warning", f"No columns found for table '{table_name}'.")
            return

        # Create a dialog to select and edit a column
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton, QLineEdit, QFormLayout

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Edit Column in {table_name}")
        dialog.resize(400, 200)

        layout = QVBoxLayout(dialog)

        # Column selection
        form_layout = QFormLayout()

        column_combo = QComboBox()
        for col in columns_info:
            column_combo.addItem(col['name'])
        form_layout.addRow("Select Column:", column_combo)

        new_name_input = QLineEdit()
        form_layout.addRow("New Name:", new_name_input)

        new_type_input = QLineEdit()
        form_layout.addRow("New Type:", new_type_input)

        layout.addLayout(form_layout)

        # Buttons
        button_layout = QHBoxLayout()

        ok_btn = QPushButton("Modify Column")
        ok_btn.clicked.connect(dialog.accept)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(dialog.reject)

        button_layout.addWidget(ok_btn)
        button_layout.addWidget(cancel_btn)

        layout.addLayout(button_layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            selected_column = column_combo.currentText()
            new_name = new_name_input.text().strip()
            new_type = new_type_input.text().strip()

            if not new_name and not new_type:
                QMessageBox.warning(self, "Warning", "Please specify at least a new name or a new type for the column.")
                return

            try:
                # Build ALTER TABLE query based on database type
                if self.db_manager.db_type == 'postgresql':
                    alter_parts = []

                    if new_name and new_name != selected_column:
                        alter_parts.append(f"RENAME COLUMN {selected_column} TO {new_name}")

                    if new_type:
                        alter_parts.append(f"ALTER COLUMN {new_name or selected_column} TYPE {new_type}")

                    if alter_parts:
                        alter_query = f"ALTER TABLE {table_name} " + ", ".join(alter_parts) + ";"
                        self.db_manager.execute_query(alter_query)

                        QMessageBox.information(self, "Success", f"Column '{selected_column}' modified successfully!")
                        self.refresh_current_table()  # Refresh the table view
                    else:
                        QMessageBox.information(self, "Info", "No changes were made to the column.")
                else:
                    # For PostgreSQL, show success message
                    QMessageBox.information(self, "Success", self.get_translation("Column modified successfully."))
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to modify column: {str(e)}")

    def add_row(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        selected_items = self.tables_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", self.get_translation("Please select a table first."))
            return

        table_name = selected_items[0].text()

        # Get table columns
        try:
            columns = self.db_manager.get_table_info(table_name)
            if not columns:
                QMessageBox.warning(self, "Warning", f"Table '{table_name}' has no columns.")
                return

            # Get the primary key column for potential undo operation
            pk_columns = [col for col in columns if col['primary_key']]
            pk_column = pk_columns[0] if pk_columns else None

            # For simplicity, we'll add a row with NULL values for non-primary key columns
            # In a real application, you might want to show a form for data entry
            non_pk_columns = [col['name'] for col in columns if not col['primary_key']]

            if non_pk_columns:
                placeholders = ", ".join(["?" for _ in non_pk_columns])
                columns_str = ", ".join(non_pk_columns)
                query = f"INSERT INTO {table_name} ({columns_str}) VALUES ({placeholders})"
                # Insert NULL values for the new row
                null_values = [None for _ in non_pk_columns]

                # Create undo action before executing the query
                def undo_func():
                    # Get the ID of the newly inserted row to delete it
                    if pk_column:
                        get_id_query = f"SELECT MAX({pk_column['name']}) FROM {table_name}"
                        result = self.db_manager.execute_query(get_id_query)
                        if result and result[0][0] is not None:
                            new_id = result[0][0]
                            delete_query = f"DELETE FROM {table_name} WHERE {pk_column['name']} = ?"
                            self.db_manager.execute_query(delete_query, (new_id,))
                            self.refresh_current_table()

                def redo_func():
                    # Rebuild query and values based on current table structure
                    columns = self.db_manager.get_table_info(table_name)
                    non_pk_columns = [col['name'] for col in columns if not col['primary_key']]

                    if non_pk_columns:
                        placeholders = ", ".join(["?" for _ in non_pk_columns])
                        columns_str = ", ".join(non_pk_columns)
                        query_to_execute = f"INSERT INTO {table_name} ({columns_str}) VALUES ({placeholders})"
                        null_values_to_use = [None for _ in non_pk_columns]
                        self.db_manager.execute_query(query_to_execute, null_values_to_use)
                    else:
                        # If there are only primary key columns, insert default values
                        query_to_execute = f"INSERT INTO {table_name} DEFAULT VALUES"
                        self.db_manager.execute_query(query_to_execute)

                    self.refresh_current_table()

                # Execute the insert
                redo_func()

                # Add to action history for undo/redo
                if self.backup_manager:
                    self.backup_manager.create_db_operation_action(
                        operation_name="add_row",
                        undo_func=undo_func,
                        redo_func=redo_func,
                        description=f"Add row to {table_name}"
                    )
                    self.update_undo_redo_states()

            else:
                # If there are only primary key columns, we might not need to insert anything
                # or use DEFAULT values
                pk_columns = [col['name'] for col in columns if col['primary_key']]
                if pk_columns:
                    # If there's an auto-increment primary key, just insert default/NULL
                    query = f"INSERT INTO {table_name} DEFAULT VALUES"

                    # Create undo action for DEFAULT VALUES insertion
                    def undo_func():
                        if pk_columns:
                            get_id_query = f"SELECT MAX({pk_columns[0]}) FROM {table_name}"
                            result = self.db_manager.execute_query(get_id_query)
                            if result and result[0][0] is not None:
                                new_id = result[0][0]
                                delete_query = f"DELETE FROM {table_name} WHERE {pk_columns[0]} = ?"
                                self.db_manager.execute_query(delete_query, (new_id,))
                                self.refresh_current_table()

                    def redo_func():
                        self.db_manager.execute_query(query)
                        self.refresh_current_table()

                    # Execute the insert
                    redo_func()

                    # Add to action history for undo/redo
                    if self.backup_manager:
                        self.backup_manager.create_db_operation_action(
                            operation_name="add_default_row",
                            undo_func=undo_func,
                            redo_func=redo_func,
                            description=f"Add default row to {table_name}"
                        )
                        self.update_undo_redo_states()

            self.statusBar().showMessage(self.get_translation("row_added_to_table").format(table=table_name))
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to add row: {str(e)}")

    def delete_row(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        selected_items = self.table_view.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", self.get_translation("Please select a row to delete."))
            return

        # Get the selected row index
        row_index = selected_items[0].row()

        selected_table_items = self.tables_list.selectedItems()
        if not selected_table_items:
            QMessageBox.warning(self, "Warning", self.get_translation("Please select a table first."))
            return

        table_name = selected_table_items[0].text()

        # Get the primary key value for the selected row
        try:
            columns = self.db_manager.get_table_info(table_name)
            pk_columns = [col for col in columns if col['primary_key']]

            if not pk_columns:
                QMessageBox.critical(self, "Error", self.get_translation("Cannot delete row: table has no primary key."))
                return

            # Get the primary key column name
            pk_column_name = pk_columns[0]['name']  # Use the first primary key column

            # Get the ID of the row to delete
            # Find the position of the primary key column in the table view
            all_columns = [desc[0] for desc in self.db_manager.cursor.description]
            pk_col_index = None
            for i, col_name in enumerate(all_columns):
                if col_name == pk_column_name:
                    pk_col_index = i
                    break

            if pk_col_index is not None:
                pk_value = self.table_view.item(row_index, pk_col_index).text()

                reply = QMessageBox.question(
                    self,
                    "Confirm Delete",
                    f"Are you sure you want to delete row with {pk_column_name}={pk_value}?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )

                if reply == QMessageBox.StandardButton.Yes:
                    # Get the entire row data for potential undo operation
                    query_get_row = f"SELECT * FROM {table_name} WHERE {pk_column_name} = ?"
                    row_data = self.db_manager.execute_query(query_get_row, (pk_value,))

                    if row_data and len(row_data) > 0:
                        row_values = row_data[0]

                        def undo_func():
                            # Reconstruct the INSERT query to restore the row
                            columns_info = [col['name'] for col in columns]
                            placeholders = ", ".join(["?" for _ in columns_info])
                            insert_query = f"INSERT INTO {table_name} ({', '.join(columns_info)}) VALUES ({placeholders})"
                            self.db_manager.execute_query(insert_query, row_values)
                            self.refresh_current_table()

                        def redo_func():
                            delete_query = f"DELETE FROM {table_name} WHERE {pk_column_name} = ?"
                            self.db_manager.execute_query(delete_query, (pk_value,))
                            self.refresh_current_table()

                        # Execute the delete
                        redo_func()

                        # Add to action history for undo/redo
                        if self.backup_manager:
                            self.backup_manager.create_db_operation_action(
                                operation_name="delete_row",
                                undo_func=undo_func,
                                redo_func=redo_func,
                                description=f"Delete row from {table_name} where {pk_column_name} = {pk_value}"
                            )
                            self.update_undo_redo_states()

                        self.statusBar().showMessage(self.get_translation("row_deleted_from_table").format(table=table_name))
                    else:
                        QMessageBox.warning(self, "Warning", "Row data not found for undo operation.")
            else:
                QMessageBox.critical(self, "Error", f"Primary key column {pk_column_name} not found in view")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to delete row: {str(e)}")

    def closeEvent(self, event):
        """Handle application close event to save important data and confirm exit"""
        # First, check if there are uncommitted transactions
        if self.db_manager and self.db_manager.in_transaction:
            reply = QMessageBox.question(
                self,
                "Uncommitted Transaction",
                "There is an uncommitted transaction. Do you want to commit before closing?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
            )

            if reply == QMessageBox.StandardButton.Yes:
                self.db_manager.commit_transaction()
            elif reply == QMessageBox.StandardButton.Cancel:
                event.ignore()
                return

        # Show exit confirmation dialog
        if self.user_mode == 'user':
            exit_message = "Are you sure you want to exit the application? (User Mode)"
        else:
            exit_message = "Are you sure you want to exit the application? (Administrator Mode)"

        # Create a custom confirmation dialog to allow for specific layout control
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton

        # Create a custom dialog to control margins
        confirm_dialog = QDialog(self)
        confirm_dialog.setWindowTitle("Confirm Exit")
        confirm_dialog.setModal(True)
        confirm_dialog.setFixedSize(350, 150)

        layout = QVBoxLayout()
        # Set left margin to 10 pixels as requested
        layout.setContentsMargins(10, 10, 10, 10)  # [left, top, right, bottom]
        layout.setSpacing(10)

        # Message label
        label = QLabel(exit_message)
        label.setWordWrap(True)
        layout.addWidget(label)

        # Buttons layout
        button_layout = QHBoxLayout()

        # Add stretch to push buttons to the right
        button_layout.addStretch()

        # No button
        no_btn = QPushButton("No")
        no_btn.setObjectName("cancel")  # Red on hover for cancel
        no_btn.clicked.connect(confirm_dialog.reject)

        # Yes button
        yes_btn = QPushButton("Yes")
        yes_btn.setObjectName("yes")  # Green on hover for yes
        yes_btn.clicked.connect(confirm_dialog.accept)  # Connect to accept

        button_layout.addWidget(no_btn)
        button_layout.addWidget(yes_btn)

        layout.addLayout(button_layout)
        confirm_dialog.setLayout(layout)

        reply = confirm_dialog.exec()

        if reply == QDialog.DialogCode.Accepted:
            # Save the current state of the database if needed
            if self.backup_manager and self.db_manager:
                # Create an auto backup on close if enabled
                try:
                    # Assuming we have the database file path stored somewhere
                    # In a real application, you'd have the actual db path
                    pass
                except:
                    # If backup fails, still allow the application to close
                    pass

            # Save query history
            if hasattr(self, 'query_history_manager'):
                try:
                    self.query_history_manager.save_queries()
                except:
                    pass

            # Accept the close event
            event.accept()
        else:
            # User chose not to exit
            event.ignore()

    def edit_row(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        selected_items = self.table_view.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", "Please select a row to edit.")
            return

        # Get the selected row index
        row_index = selected_items[0].row()

        # Get the table name from the selected table in the left panel
        selected_table_items = self.tables_list.selectedItems()
        if not selected_table_items:
            QMessageBox.warning(self, "Warning", "Please select a table first.")
            return

        table_name = selected_table_items[0].text()

        # Get table columns
        try:
            columns = self.db_manager.get_table_info(table_name)
            if not columns:
                QMessageBox.warning(self, "Warning", f"Table '{table_name}' has no columns.")
                return

            # Get primary key information to identify the specific row
            pk_columns = [col for col in columns if col['primary_key']]
            if not pk_columns:
                QMessageBox.warning(self, "Warning", f"Table '{table_name}' has no primary key. Cannot edit rows without a primary key.")
                return

            pk_column = pk_columns[0]  # Use first primary key
            pk_column_index = None

            # Find the index of the primary key column in the table view
            for i in range(self.table_view.columnCount()):
                if self.table_view.horizontalHeaderItem(i).text() == pk_column['name']:
                    pk_column_index = i
                    break

            if pk_column_index is None:
                QMessageBox.critical(self, "Error", f"Could not find primary key column '{pk_column['name']}' in view.")
                return

            # Get the primary key value for the selected row
            pk_value = self.table_view.item(row_index, pk_column_index).text()

            # Get the current values for this row from the database
            query = f"SELECT * FROM {table_name} WHERE {pk_column['name']} = ?"
            result = self.db_manager.execute_query(query, (pk_value,))

            if not result or len(result) == 0:
                QMessageBox.critical(self, "Error", "Could not retrieve row data.")
                return

            current_values = result[0]  # Assuming single row result

            # Create and show the UpdateEntityDialog
            # Create a simplified column info structure for the dialog
            simplified_columns = []
            for i, col in enumerate(columns):
                col_info = col.copy()
                # Add the current value for the dialog
                col_info['current_value'] = str(current_values[i]) if current_values[i] is not None else ""
                simplified_columns.append(col_info)

            dialog = UpdateEntityDialog(self, table_name, simplified_columns, f"{pk_column['name']} = {pk_value}")

            if dialog.exec() == QDialog.DialogCode.Accepted:
                updates = dialog.get_updates()
                if updates:
                    # Build the UPDATE query
                    set_clause = ", ".join([f"{col} = ?" for col in updates.keys()])
                    query = f"UPDATE {table_name} SET {set_clause} WHERE {pk_column['name']} = ?"

                    # Create undo action before executing the query
                    def undo_func():
                        # Revert to the original values
                        original_set_clause = ", ".join([f"{col['name']} = ?" for col in columns if col['name'] in [k for k in updates.keys()] and col['name'] != pk_column['name']])
                        if original_set_clause:
                            undo_query = f"UPDATE {table_name} SET {original_set_clause} WHERE {pk_column['name']} = ?"
                            # Get original values for the columns being updated
                            original_values = [result[0][i] for i, col in enumerate(columns) if col['name'] in updates.keys()]
                            self.db_manager.execute_query(undo_query, original_values + [pk_value])
                            self.refresh_current_table()

                    def redo_func():
                        # Execute the update query with the new values
                        self.db_manager.execute_query(query, tuple(updates.values()) + (pk_value,))
                        self.refresh_current_table()

                    # Execute the update
                    redo_func()

                    # Add to action history for undo/redo
                    if self.backup_manager:
                        self.backup_manager.create_db_operation_action(
                            operation_name="update_row",
                            undo_func=undo_func,
                            redo_func=redo_func,
                            description=f"Update row in {table_name} where {pk_column['name']} = {pk_value}"
                        )
                        self.update_undo_redo_states()

                    self.statusBar().showMessage(f"Row updated in {table_name}")
                else:
                    # No updates provided, inform user
                    pass
            else:
                # User cancelled the dialog
                pass

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to edit row: {str(e)}")

    def execute_predefined_query(self, query):
        """Execute a predefined query and display results in the query results tab"""
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        try:
            # Switch to the Query Editor tab to show results
            self.tab_widget.setCurrentIndex(1)  # Index 1 is the query editor tab

            # Display the query in the input field
            self.query_input.setPlainText(query)

            # Execute the query
            result = self.db_manager.execute_query(query)

            if result is not None:  # SELECT query
                if result:
                    # Get column names
                    columns = [desc[0] for desc in self.db_manager.cursor.description]

                    # Set up the results table
                    self.query_results.setRowCount(len(result))
                    self.query_results.setColumnCount(len(columns))
                    self.query_results.setHorizontalHeaderLabels(columns)

                    # Populate the table with color coding for data types
                    for i, row in enumerate(result):
                        for j, value in enumerate(row):
                            item = QTableWidgetItem(str(value))

                            # Determine data type and apply appropriate styling
                            value_str = str(value)
                            if value_str.isdigit() or (value_str.startswith('-') and value_str[1:].isdigit()) or self.is_numeric(value_str):
                                # Numeric data - greenish
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_NUMERIC)
                                item.setBackground(QColor(COLOR_NUMERIC_BG))
                                item.setForeground(QColor(COLOR_NUMERIC_FG))
                            elif self.is_date_or_time(value_str):
                                # Date/time data - orange
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_DATE)
                                item.setBackground(QColor(COLOR_DATE_BG))
                                item.setForeground(QColor(COLOR_DATE_FG))
                            elif value_str.lower() in ['true', 'false', '1', '0', 'yes', 'no']:
                                # Boolean data - teal
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_BOOLEAN)
                                item.setBackground(QColor(COLOR_BOOLEAN_BG))
                                item.setForeground(QColor(COLOR_BOOLEAN_FG))
                            else:
                                # Text data - blueish
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_TEXT)
                                item.setBackground(QColor(COLOR_TEXT_BG))
                                item.setForeground(QColor(COLOR_TEXT_FG))

                            self.query_results.setItem(i, j, item)

                    # Resize columns to fit content
                    self.query_results.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
                else:
                    # No results but query executed successfully
                    self.query_results.setRowCount(0)
                    self.query_results.setColumnCount(0)

                self.statusBar().showMessage(self.get_translation("predefined_query_executed_successfully_rows").format(count=len(result) if result else 0))
            else:  # Non-SELECT query
                self.query_results.setRowCount(0)
                self.query_results.setColumnCount(0)
                self.statusBar().showMessage(self.get_translation("predefined_query_executed_successfully"))

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error executing predefined query: {str(e)}")

    def execute_query(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        query = self.query_input.toPlainText()
        if not query.strip():
            QMessageBox.warning(self, "Warning", self.get_translation("Please enter a query to execute."))
            return

        # Check user mode - if in user mode, restrict data modification queries
        if self.user_mode == 'user':
            query_upper = query.strip().upper()
            if query_upper.startswith(('INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP', 'ALTER')):
                QMessageBox.warning(
                    self,
                    "Access Restricted",
                    "You are in User Mode (Read-only). Data modification queries are not allowed."
                )
                return

        # For non-SELECT queries (INSERT, UPDATE, DELETE), we want to track changes for undo/redo
        query_upper = query.strip().upper()

        if query_upper.startswith('INSERT') or query_upper.startswith('UPDATE') or query_upper.startswith('DELETE'):
            # For data modification queries, we'll track the previous state for potential undo
            affected_records = None

            # If this is an UPDATE or DELETE, capture data that might be modified
            table_name = self.extract_table_name_from_query(query)

            if table_name and (query_upper.startswith('UPDATE') or query_upper.startswith('DELETE')):
                try:
                    # Instead of selecting all data, parse the WHERE clause to get specific records
                    # For now, we'll just note that this is an update/delete operation
                    # In a production system, we would parse the WHERE clause to get specific records
                    affected_records = "Records matching WHERE clause"
                except Exception as parse_error:
                    self.logger.error(f"Error parsing query for affected records: {str(parse_error)}")
                    affected_records = None  # If we can't parse the query, continue anyway

            try:
                result = self.db_manager.execute_query(query)

                # For data modification queries, set up undo/redo actions
                def undo_func():
                    if query_upper.startswith('INSERT'):
                        # For INSERT, we can try to delete the last inserted record based on PK
                        try:
                            columns_info = self.db_manager.get_table_info(table_name)
                            pk_columns = [col for col in columns_info if col['primary_key']]
                            if pk_columns:
                                pk_col_name = pk_columns[0]['name']
                                get_id_query = f"SELECT MAX({pk_col_name}) FROM {table_name}"
                                result_id = self.db_manager.execute_query(get_id_query)
                                if result_id and result_id[0][0] is not None:
                                    delete_query = f"DELETE FROM {table_name} WHERE {pk_col_name} = %s"
                                    self.db_manager.execute_query(delete_query, (result_id[0][0],))
                                    self.refresh_current_table()
                        except Exception as undo_error:
                            self.logger.error(f"Error during undo operation: {str(undo_error)}")
                            QMessageBox.warning(self, "Undo Error", f"Could not perform undo operation: {str(undo_error)}")
                    elif query_upper.startswith('UPDATE'):
                        # For UPDATE, we would need to store the original values to restore them
                        # This is a simplified approach - in a real system, we'd need to store the original values
                        pass
                    elif query_upper.startswith('DELETE'):
                        # For DELETE, we would need to restore the deleted records
                        # This is a simplified approach - in a real system, we'd need to store the deleted records
                        pass

                def redo_func():
                    try:
                        self.db_manager.execute_query(query)
                        self.refresh_current_table()
                    except Exception as redo_error:
                        self.logger.error(f"Error during redo operation: {str(redo_error)}")
                        QMessageBox.critical(self, "Redo Error", f"Could not perform redo operation: {str(redo_error)}")

                # Add to action history for undo/redo
                if self.backup_manager:
                    try:
                        self.backup_manager.create_db_operation_action(
                            operation_name="execute_query",
                            undo_func=undo_func,
                            redo_func=redo_func,
                            description=f"Execute query: {query[:50]}{'...' if len(query) > 50 else ''}"
                        )
                        self.update_undo_redo_states()
                    except Exception as action_error:
                        self.logger.error(f"Error creating undo/redo action: {str(action_error)}")

                self.query_results.setRowCount(0)
                self.query_results.setColumnCount(0)
                self.statusBar().showMessage(self.get_translation("query_executed_successfully"))

            except Exception as e:
                error_msg = f"Error executing query: {str(e)}"
                self.logger.error(error_msg)
                QMessageBox.critical(self, "Error", error_msg)
        else:
            # For SELECT queries, just execute normally without undo/redo tracking
            try:
                result = self.db_manager.execute_query(query)

                if result is not None:  # SELECT query
                    if result:
                        # Get column names
                        columns = [desc[0] for desc in self.db_manager.cursor.description]

                        # Set up the results table
                        self.query_results.setRowCount(len(result))
                        self.query_results.setColumnCount(len(columns))
                        self.query_results.setHorizontalHeaderLabels(columns)

                        # Populate the table
                        for i, row in enumerate(result):
                            for j, value in enumerate(row):
                                item = QTableWidgetItem(str(value) if value is not None else "")
                                # Apply highlighting based on data type
                                self.apply_data_type_highlighting(item, str(value) if value is not None else "")
                                self.query_results.setItem(i, j, item)

                        # Resize columns to fit content
                        self.query_results.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
                    else:
                        # No results but query executed successfully
                        self.query_results.setRowCount(0)
                        self.query_results.setColumnCount(0)

                    self.statusBar().showMessage(self.get_translation("query_executed_successfully_rows").format(count=len(result) if result else 0))
                else:  # Non-SELECT query that's not INSERT/UPDATE/DELETE
                    self.query_results.setRowCount(0)
                    self.query_results.setColumnCount(0)
                    self.statusBar().showMessage(self.get_translation("query_executed_successfully"))

                # Add the query to history with a descriptive name
                query_description = f"Manual query: {query.strip()[:50]}{'...' if len(query.strip()) > 50 else ''}"
                self.query_history_manager.add_query(query, query_description)
                self.refresh_query_history()

            except Exception as e:
                error_msg = f"Error executing query: {str(e)}"
                self.logger.error(error_msg)
                QMessageBox.critical(self, "Error", error_msg)

    def extract_table_name_from_query(self, query: str) -> str:
        """Extract table name from INSERT, UPDATE, or DELETE query"""
        query_upper = query.strip().upper()

        if query_upper.startswith('INSERT'):
            # For INSERT: "INSERT INTO table_name ..."
            parts = query.split()
            try:
                into_idx = next(i for i, part in enumerate(parts) if part.upper() == 'INTO')
                if into_idx + 1 < len(parts):
                    return parts[into_idx + 1]
            except:
                pass
        elif query_upper.startswith('UPDATE'):
            # For UPDATE: "UPDATE table_name ..."
            parts = query.split()
            if len(parts) > 1:
                return parts[1]
        elif query_upper.startswith('DELETE'):
            # For DELETE: "DELETE FROM table_name ..."
            parts = query.split()
            try:
                from_idx = next(i for i, part in enumerate(parts) if part.upper() == 'FROM')
                if from_idx + 1 < len(parts):
                    return parts[from_idx + 1]
            except:
                pass

        return None

    def execute_transaction_query(self):
        """Execute multiple queries as a transaction from the query editor"""
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        # Check user mode - if in user mode, restrict data modification queries
        if self.user_mode == 'user':
            query_text = self.query_input.toPlainText()
            query_upper = query_text.strip().upper()
            if any(mod_query in query_upper for mod_query in ['INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP', 'ALTER']):
                QMessageBox.warning(
                    self,
                    "Access Restricted",
                    "You are in User Mode (Read-only). Data modification queries are not allowed."
                )
                return

        query_text = self.query_input.toPlainText()
        if not query_text.strip():
            QMessageBox.warning(self, "Warning", self.get_translation("Please enter queries to execute as a transaction."))
            return

        # Split the query text by semicolons to separate individual queries
        # This is a basic approach - in a real system you might want more sophisticated SQL parsing
        queries = [q.strip() for q in query_text.split(';') if q.strip()]

        if not queries:
            QMessageBox.warning(self, "Warning", self.get_translation("No valid queries found in the input."))
            return

        reply = QMessageBox.question(
            self,
            "Confirm Transaction Execution",
            f"Execute {len(queries)} queries as a single transaction?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                results = self.db_manager.execute_transaction(queries)
                QMessageBox.information(self, "Transaction", f"Transaction executed successfully. {len(results)} queries processed.")
                self.statusBar().showMessage(self.get_translation("transaction_completed_operations").format(count=len(results)))

                # Refresh the current table view if it's a data modification transaction
                if any('INSERT' in q.upper() or 'UPDATE' in q.upper() or 'DELETE' in q.upper() for q in queries):
                    self.refresh_current_table()

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Transaction failed and was rolled back: {str(e)}")

    def save_query(self):
        query = self.query_input.toPlainText()
        if not query.strip():
            QMessageBox.warning(self, "Warning", self.get_translation("Please enter a query to save."))
            return

        description, ok = QInputDialog.getText(self, self.get_translation("Save Query"), self.get_translation("Enter description for this query:"))
        if ok:
            self.query_history_manager.add_query(query, description)
            self.refresh_query_history()
            self.statusBar().showMessage(self.get_translation("query_saved_to_history"))

    def refresh_query_history(self):
        queries = self.query_history_manager.get_queries()

        self.history_list.setRowCount(len(queries))
        for i, query_entry in enumerate(queries):
            self.history_list.setItem(i, 0, QTableWidgetItem(query_entry['description']))
            self.history_list.setItem(i, 1, QTableWidgetItem(query_entry['query']))
            self.history_list.setItem(i, 2, QTableWidgetItem(query_entry['timestamp']))

    def load_query_from_history(self, row, column):
        # Load the query from the selected row into the query input
        query_item = self.history_list.item(row, 1)  # Column 1 contains the query
        if query_item:
            self.query_input.setPlainText(query_item.text())
            # Switch to the query editor tab
            self.tab_widget.setCurrentIndex(1)  # Index 1 is the query editor tab

    def load_selected_query(self):
        selected_items = self.history_list.selectedItems()
        if selected_items:
            row = selected_items[0].row()
            query_item = self.history_list.item(row, 1)  # Column 1 contains the query
            if query_item:
                self.query_input.setPlainText(query_item.text())
                # Switch to the query editor tab
                self.tab_widget.setCurrentIndex(1)  # Index 1 is the query editor tab

    def delete_query_from_history(self):
        selected_items = self.history_list.selectedItems()
        if selected_items:
            row = selected_items[0].row()
            reply = QMessageBox.question(
                self,
                "Confirm Delete",
                "Are you sure you want to delete this query from history?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                # Remove from the in-memory list
                self.query_history_manager.queries.pop(row)
                # Save to file
                self.query_history_manager.save_queries()
                # Refresh the list
                self.refresh_query_history()

    def create_backup(self):
        if not self.backup_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Create Backup",
            f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON Files (*.json);;All Files (*)"
        )

        if file_path:
            if self.backup_manager.create_backup(file_path):
                self.statusBar().showMessage(self.get_translation("backup_created_successfully").format(path=file_path))
                QMessageBox.information(self, "Success", f"Backup created successfully: {file_path}")
            else:
                QMessageBox.critical(self, "Error", self.get_translation("backup_creation_failed"))

    def restore_backup(self):
        if not self.backup_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Restore from Backup",
            "",
            "JSON Files (*.json);;All Files (*)"
        )

        if file_path:
            reply = QMessageBox.question(
                self,
                "Confirm Restore",
                "This will replace all existing data. Are you sure you want to restore?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                # To enable undo, we should create a backup of the current state first
                current_data_backup = self.create_temp_backup()

                if current_data_backup:
                    def undo_func():
                        # Restore from the backup created before the restore operation
                        self.backup_manager.restore_backup(current_data_backup)
                        self.refresh_tables()
                        self.statusBar().showMessage("Database restore operation undone")

                    def redo_func():
                        # Perform the original restore operation
                        self.backup_manager.restore_backup(file_path)
                        self.refresh_tables()
                        self.statusBar().showMessage(self.get_translation("database_restored_successfully").format(path=file_path))

                    # Execute the restore
                    if self.backup_manager.restore_backup(file_path):
                        self.refresh_tables()

                        # Add to action history for undo/redo
                        if self.backup_manager:
                            self.backup_manager.create_db_operation_action(
                                operation_name="restore_backup",
                                undo_func=undo_func,
                                redo_func=redo_func,
                                description=f"Restore database from {file_path}"
                            )
                            self.update_undo_redo_states()

                        self.statusBar().showMessage(self.get_translation("database_restored_successfully").format(path=file_path))
                        QMessageBox.information(self, "Success", f"Database restored successfully: {file_path}")
                    else:
                        # Clean up the temporary backup if restore failed
                        import os
                        if os.path.exists(current_data_backup):
                            os.remove(current_data_backup)
                        QMessageBox.critical(self, "Error", self.get_translation("restore_failed"))
                else:
                    QMessageBox.critical(self, "Error", "Could not create backup before restore operation - restore aborted.")

    def create_temp_backup(self):
        """Create a temporary backup of the current database state for potential undo operations"""
        import tempfile
        import os
        from datetime import datetime

        # Create a temporary file for the backup
        temp_dir = os.path.join("backups", "temp")
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        temp_backup_path = os.path.join(temp_dir, f"temp_backup_{timestamp}.json")

        try:
            if self.backup_manager.create_backup(temp_backup_path):
                return temp_backup_path
            else:
                return None
        except Exception as e:
            print(f"Error creating temporary backup: {e}")
            return None

    def export_table_to_excel(self):
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        selected_items = self.tables_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", self.get_translation("Please select a table to export."))
            return

        table_name = selected_items[0].text()

        # Get the table data for export
        query = f"SELECT * FROM {table_name}"
        data = self.db_manager.execute_query(query)

        if not data:
            QMessageBox.warning(self, "Warning", f"Table '{table_name}' has no data to export.")
            return

        # Get column names
        columns = [desc[0] for desc in self.db_manager.cursor.description]

        # Initialize export utilities if not already done
        if not hasattr(self, 'export_utils'):
            self.export_utils = ExportUtils(self)

        # Ask user for export format
        from PyQt6.QtWidgets import QInputDialog
        format_options = [
            "Excel (.xlsx)",
            "Text (.txt)",
            "Markdown (.md)",
            "SQL (.sql)"
        ]

        format_choice, ok = QInputDialog.getItem(
            self,
            "Export Format",
            "Choose export format:",
            format_options,
            0,
            False
        )

        if not ok:
            return  # User canceled

        # Create a temporary table widget to hold the data for export
        temp_table = QTableWidget()
        temp_table.setRowCount(len(data) if data else 0)
        temp_table.setColumnCount(len(columns))
        temp_table.setHorizontalHeaderLabels(columns)

        # Populate the temporary table with data
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_value in enumerate(row_data):
                temp_table.setItem(row_idx, col_idx, QTableWidgetItem(str(cell_value)))

        if format_choice == "Excel (.xlsx)":
            # Original Excel export functionality
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export to...",
                f"{table_name}.xlsx",
                "Excel Files (*.xlsx);;All Files (*)"
            )

            if file_path:
                try:
                    # Create Excel workbook
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = table_name

                    # Add headers
                    for col_num, column_title in enumerate(columns, 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.value = column_title
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                    # Add data
                    for row_num, row_data in enumerate(data, 2):
                        for col_num, cell_value in enumerate(row_data, 1):
                            ws.cell(row=row_num, column=col_num, value=str(cell_value))

                    # Adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Limit max width
                        ws.column_dimensions[column_letter].width = adjusted_width

                    # Save workbook
                    wb.save(file_path)

                    self.statusBar().showMessage(self.get_translation("table_exported_successfully").format(table=table_name, path=file_path))
                    QMessageBox.information(self, "Success", f"Table '{table_name}' exported to {file_path}")

                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to export table: {str(e)}")

        elif format_choice == "Text (.txt)":
            self.export_utils.export_to_txt(temp_table, f"Table: {table_name}")

        elif format_choice == "Markdown (.md)":
            self.export_utils.export_to_md(temp_table, f"Table: {table_name}")


        elif format_choice == "SQL (.sql)":
            self.export_utils.export_to_sql(temp_table, query, table_name)

    def export_query_results(self):
        if self.query_results.rowCount() == 0 or self.query_results.columnCount() == 0:
            QMessageBox.warning(self, "Warning", self.get_translation("No query results to export."))
            return

        # Initialize export utilities if not already done
        if not hasattr(self, 'export_utils'):
            self.export_utils = ExportUtils(self)

        # Ask user for export format
        from PyQt6.QtWidgets import QInputDialog
        format_options = [
            "Excel (.xlsx)",
            "Text (.txt)",
            "Markdown (.md)",
            "SQL (.sql)"
        ]

        format_choice, ok = QInputDialog.getItem(
            self,
            "Export Format",
            "Choose export format:",
            format_options,
            0,
            False
        )

        if not ok:
            return  # User canceled

        # Get current query from query_input
        current_query = self.query_input.toPlainText().strip()

        if format_choice == "Excel (.xlsx)":
            # Original Excel export functionality
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export to...",
                f"query_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx);;All Files (*)"
            )

            if file_path:
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Query Results"

                    # Add headers from table widget
                    for col in range(self.query_results.columnCount()):
                        header_item = self.query_results.horizontalHeaderItem(col)
                        header_text = header_item.text() if header_item else f"Column {col+1}"
                        cell = ws.cell(row=1, column=col+1)
                        cell.value = header_text
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                    # Add data from table widget
                    for row in range(self.query_results.rowCount()):
                        for col in range(self.query_results.columnCount()):
                            item = self.query_results.item(row, col)
                            cell_value = item.text() if item else ""
                            ws.cell(row=row+2, column=col+1, value=str(cell_value))

                    # Adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Limit max width
                        ws.column_dimensions[column_letter].width = adjusted_width

                    # Save workbook
                    wb.save(file_path)

                    self.statusBar().showMessage(self.get_translation("query_results_exported_successfully").format(path=file_path))
                    QMessageBox.information(self, "Success", f"Query results exported to {file_path}")

                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to export query results: {str(e)}")

        elif format_choice == "Text (.txt)":
            self.export_utils.export_to_txt(self.query_results, "Query Results")

        elif format_choice == "Markdown (.md)":
            self.export_utils.export_to_md(self.query_results, "Query Results")


        elif format_choice == "SQL (.sql)":
            # Use current query or default name
            query_title = current_query[:50] + "..." if len(current_query) > 50 else current_query
            if not query_title.strip():
                query_title = "SELECT * FROM table"  # Default query name
            self.export_utils.export_to_sql(self.query_results, current_query, "exported_data")

    def change_language(self, language_code):
        """Change the application language"""
        # Update radio button states
        self.lang_ru_action.setChecked(language_code == 'ru')
        self.lang_be_action.setChecked(language_code == 'be')
        self.lang_en_action.setChecked(language_code == 'en')

        self.current_language = language_code
        self.translation_manager.set_language(language_code)

        # Store the selected language
        self.selected_language = language_code

        # Refresh the entire UI with new language
        self.refresh_ui_language()

    def refresh_ui_language(self):
        """Refresh all UI elements with the current language"""
        # Refresh main window title
        self.setWindowTitle(self.get_translation("Database Management System"))

        # Refresh tab names
        self.tab_widget.setTabText(0, self.get_translation("table_view_tab"))
        self.tab_widget.setTabText(1, self.get_translation("query_editor_tab"))
        self.tab_widget.setTabText(2, self.get_translation("query_history_tab"))
        self.tab_widget.setTabText(3, self.get_translation("special_queries_tab"))
        self.tab_widget.setTabText(4, self.get_translation("entity_operations_tab"))
        self.tab_widget.setTabText(5, self.get_translation("select_queries_tab"))
        self.tab_widget.setTabText(6, self.get_translation("join_queries_tab"))
        self.tab_widget.setTabText(7, self.get_translation("aggregate_queries_tab"))

        # Refresh table view tab controls
        self.refresh_data_btn.setText(self.get_translation("Refresh Data"))
        self.add_row_btn.setText(self.get_translation("Add Row"))
        self.delete_row_btn.setText(self.get_translation("Delete Row"))
        self.edit_row_btn.setText(self.get_translation("Edit Row"))

        # Refresh query editor tab controls
        self.query_input.setPlaceholderText(self.get_translation("Enter SQL query here..."))
        self.execute_btn.setText(self.get_translation("Execute Query"))
        self.execute_transaction_btn.setText(self.get_translation("Execute as Transaction"))
        self.save_query_btn.setText(self.get_translation("Save Query"))
        self.export_query_btn.setText(self.get_translation("Export to..."))

        # Refresh query history tab controls
        self.history_list.setHorizontalHeaderLabels([
            self.get_translation("Description"),
            self.get_translation("Query"),
            self.get_translation("Timestamp")
        ])
        self.load_query_btn.setText(self.get_translation("Load Query"))
        self.delete_history_btn.setText(self.get_translation("Delete Query"))

        # Refresh special queries tab
        self.update_special_queries_tab_language()

        # Refresh special queries menu items
        self.update_special_queries_menu_language()


        # Update left panel elements (table and column names)
        if hasattr(self, 'tables_list'):
            if self.tables_list.columnCount() > 0:
                self.tables_list.setHorizontalHeaderLabels([self.get_translation("Table")])

        if hasattr(self, 'columns_list'):
            if self.columns_list.columnCount() >= 5:
                self.columns_list.setHorizontalHeaderLabels([
                    self.get_translation("Name"),
                    self.get_translation("Type"),
                    self.get_translation("Not Null"),
                    self.get_translation("Default"),
                    self.get_translation("PK")
                ])

        # Update left panel button texts
        if hasattr(self, 'add_table_btn'):
            self.add_table_btn.setText(self.get_translation("Add New Table"))
        if hasattr(self, 'delete_table_btn'):
            self.delete_table_btn.setText(self.get_translation("Delete Table"))
        if hasattr(self, 'refresh_btn'):
            self.refresh_btn.setText(self.get_translation("Refresh Data"))
        if hasattr(self, 'add_column_btn'):
            self.add_column_btn.setText(self.get_translation("Add Column"))
        if hasattr(self, 'delete_column_btn'):
            self.delete_column_btn.setText(self.get_translation("Delete Column"))
        if hasattr(self, 'edit_column_btn'):
            self.edit_column_btn.setText(self.get_translation("Edit Column"))

        # Refresh select queries tab
        self.update_select_queries_tab_language()

        # Refresh join queries tab
        self.update_join_queries_tab_language()

        # Refresh aggregate queries tab
        self.update_aggregate_queries_tab_language()

        # Refresh entity operations tab
        if hasattr(self, 'entity_operations_instructions'):
            self.entity_operations_instructions.setText(
                self.get_translation("Select entity and operation to perform:")
            )

        # Refresh saved queries tab
        self.update_saved_queries_tab_language()

        # Update menu items (this will update them as needed)
        self.update_menu_language()

        # Refresh status bar message if needed
        self.statusBar().showMessage(self.get_translation("Ready"))

    def update_special_queries_tab_language(self):
        """Update special queries tab language"""
        if hasattr(self, 'special_queries_instructions'):
            self.special_queries_instructions.setText(self.get_translation("Select and execute predefined queries from lab work #5 and #6:"))
        # Note: The special queries tab now uses organized buttons instead of combo boxes
        # So we're not updating the combo box labels since they're hidden
        if hasattr(self, 'special_queries_label'):
            self.special_queries_label.setText(self.get_translation("select_query"))
        if hasattr(self, 'special_queries_params_label'):
            self.special_queries_params_label.setText(self.get_translation("parameters"))
        if hasattr(self, 'query_params_input'):
            self.query_params_input.setPlaceholderText(self.get_translation("parameters_required"))

        # Update combo box items for backward compatibility (hidden)
        if hasattr(self, 'special_queries_combo') and self.special_queries_combo.isVisible():
            current_index = self.special_queries_combo.currentIndex()
            self.special_queries_combo.clear()
            self.special_queries_combo.addItems([
                self.get_translation("get_all_trains"),
                self.get_translation("get_all_employees"),
                self.get_translation("get_scheduled_arrivals"),
                self.get_translation("get_passengers_with_tickets"),
                self.get_translation("get_employee_train_assignments"),
                self.get_translation("get_all_scheduled_trains"),
                self.get_translation("get_passengers_tickets_trains"),
                self.get_translation("get_employees_train_details"),
                self.get_translation("get_services_employee_assignments"),
                self.get_translation("get_detailed_schedule_passengers"),
                self.get_translation("get_all_passengers_by_train"),
                self.get_translation("get_employee_assignments_by_train_type"),
                self.get_translation("get_platform_utilization"),
                self.get_translation("get_tickets_by_price_range"),
                self.get_translation("get_scheduled_trains_by_date"),
                self.get_translation("get_trains_specific_type"),
                self.get_translation("get_employees_specific_position"),
                self.get_translation("get_employees_x_years_experience"),
                self.get_translation("get_passengers_specific_passport"),
                self.get_translation("get_tickets_specific_carriage"),
                self.get_translation("get_schedule_specific_date"),
                self.get_translation("get_schedule_specific_train_id")
            ])
            self.special_queries_combo.setCurrentIndex(current_index)

        # If special query buttons exist, update them too
        # The special queries have grouped buttons for lab work 5 and 6, which need to be updated
        # Since we don't have individual references to these buttons, we'll need to recreate them
        # Or at least make sure they'll be updated when the tab is redrawn
        self.update_special_queries_buttons_language()

        # Update button texts
        if hasattr(self, 'execute_special_query_btn'):
            self.execute_special_query_btn.setText(self.get_translation("execute_selected_query"))
        if hasattr(self, 'save_special_results_btn'):
            self.save_special_results_btn.setText(self.get_translation("save_results_excel"))


        # Update operation combo box
        if hasattr(self, 'entity_operation_combo'):
            current_index = self.entity_operation_combo.currentIndex()
            self.entity_operation_combo.clear()
            self.entity_operation_combo.addItems([
                self.get_translation("view_operation"),
                self.get_translation("add_operation"),
                self.get_translation("update_operation"),
                self.get_translation("delete_operation")
            ])
            self.entity_operation_combo.setCurrentIndex(current_index)

        # Update button text
        if hasattr(self, 'execute_entity_operation_btn'):
            self.execute_entity_operation_btn.setText(self.get_translation("execute_operation"))

    def update_special_queries_buttons_language(self):
        """Update the special query buttons in Lab 5 and Lab 6 groups"""
        # Since the special queries tab buttons are created dynamically and not stored with individual references,
        # we need to refresh the special queries tab entirely by recreating its content
        # This means refreshing the groups of Lab 5 and Lab 6 query buttons
        pass  # Implementation would require storing button references, which is complex
        # Instead, we'll handle this by enhancing the special queries tab recreation functionality

    def update_special_queries_menu_language(self):
        """Update special queries menu items language"""
        # Update the tooltips and text of all special query menu actions
        # These are the 15 predefined queries from special_queries_menu
        if hasattr(self, 'query1_action'):
            self.query1_action.setText(self.get_translation('get_all_trains'))
            self.query1_action.setToolTip("Get all train records with their details")
        if hasattr(self, 'query2_action'):
            self.query2_action.setText(self.get_translation('get_all_employees'))
            self.query2_action.setToolTip("Get all employees with their positions ordered by position")
        if hasattr(self, 'query3_action'):
            self.query3_action.setText(self.get_translation('get_scheduled_arrivals'))
            self.query3_action.setToolTip("Get scheduled arrivals and departures")
        if hasattr(self, 'query4_action'):
            self.query4_action.setText(self.get_translation('get_passengers_with_tickets'))
            self.query4_action.setToolTip("Get passengers with their ticket information")
        if hasattr(self, 'query5_action'):
            self.query5_action.setText(self.get_translation('get_employee_train_assignments'))
            self.query5_action.setToolTip("Get assignment of employees to trains")
        if hasattr(self, 'query6_action'):
            self.query6_action.setText(self.get_translation('get_all_scheduled_trains'))
            self.query6_action.setToolTip("Get all scheduled trains with platform information")
        if hasattr(self, 'query7_action'):
            self.query7_action.setText(self.get_translation('get_passengers_tickets_trains'))
            self.query7_action.setToolTip("Get passengers with their tickets and train information")
        if hasattr(self, 'query8_action'):
            self.query8_action.setText(self.get_translation('get_employees_train_details'))
            self.query8_action.setToolTip("Get employees assigned to specific trains with their details")
        if hasattr(self, 'query9_action'):
            self.query9_action.setText(self.get_translation('get_services_employee_assignments'))
            self.query9_action.setToolTip("Get services with assignments to employees")
        if hasattr(self, 'query10_action'):
            self.query10_action.setText(self.get_translation('get_detailed_schedule_passengers'))
            self.query10_action.setToolTip("Get detailed schedule with passenger information")
        if hasattr(self, 'query11_action'):
            self.query11_action.setText(self.get_translation("get_all_passengers_by_train"))
            self.query11_action.setToolTip("Get all passengers by train")
        if hasattr(self, 'query12_action'):
            self.query12_action.setText(self.get_translation("get_employee_assignments_by_train_type"))
            self.query12_action.setToolTip("Get employee assignments by train type")
        if hasattr(self, 'query13_action'):
            self.query13_action.setText(self.get_translation("get_platform_utilization"))
            self.query13_action.setToolTip("Get platform utilization statistics")
        if hasattr(self, 'query14_action'):
            self.query14_action.setText(self.get_translation("get_tickets_by_price_range"))
            self.query14_action.setToolTip("Get tickets within a specific price range")
        if hasattr(self, 'query15_action'):
            self.query15_action.setText(self.get_translation("get_scheduled_trains_by_date"))
            self.query15_action.setToolTip("Get scheduled trains for a specific date")

    def update_select_queries_tab_language(self):
        """Update select queries tab language"""
        if hasattr(self, 'select_queries_instructions'):
            self.select_queries_instructions.setText(
                self.get_translation("Select and execute SELECT queries for individual tables:")
            )

        if hasattr(self, 'select_table_label'):
            self.select_table_label.setText(
                self.get_translation("select_table_for_select_queries") + ":"
            )

        if hasattr(self, 'select_query_type_label'):
            self.select_query_type_label.setText(
                self.get_translation("select_query_type") + ":"
            )

        if hasattr(self, 'save_select_query_btn'):
            self.save_select_query_btn.setText(
                self.get_translation("save_query")
            )

    def update_join_queries_tab_language(self):
        """Update join queries tab language"""
        if hasattr(self, 'join_queries_instructions'):
            self.join_queries_instructions.setText(
                self.get_translation("Select and execute JOIN queries between tables:")
            )

        if hasattr(self, 'first_table_label'):
            self.first_table_label.setText(
                self.get_translation("first_table") + ":"
            )

        if hasattr(self, 'second_table_label'):
            self.second_table_label.setText(
                self.get_translation("second_table") + ":"
            )

        if hasattr(self, 'join_type_label'):
            self.join_type_label.setText(
                self.get_translation("join_type") + ":"
            )

        if hasattr(self, 'save_join_query_btn'):
            self.save_join_query_btn.setText(
                self.get_translation("save_query")
            )

    def update_aggregate_queries_tab_language(self):
        """Update aggregate queries tab language"""
        if hasattr(self, 'aggregate_queries_instructions'):
            self.aggregate_queries_instructions.setText(
                self.get_translation("Select and execute aggregate, subquery, and set operation queries:")
            )

    def update_saved_queries_tab_language(self):
        """Update saved queries tab language"""
        if hasattr(self, 'saved_queries_table'):
            if self.saved_queries_table.columnCount() >= 3:
                self.saved_queries_table.setHorizontalHeaderLabels([
                    self.get_translation("Description"),
                    self.get_translation("Query"),
                    self.get_translation("Timestamp")
                ])
        if hasattr(self, 'refresh_saved_queries_btn'):
            self.refresh_saved_queries_btn.setText(
                self.get_translation("Refresh")
            )
        if hasattr(self, 'execute_saved_query_btn'):
            self.execute_saved_query_btn.setText(
                self.get_translation("execute_selected_query")
            )
        if hasattr(self, 'delete_saved_query_btn'):
            self.delete_saved_query_btn.setText(
                self.get_translation("Delete Query")
            )

    def update_menu_language(self):
        """Update menu items language"""
        # This function is a placeholder; menu items are updated through translations when actions are triggered
        # When get_translation is called during UI setup, they automatically reflect the current language

    def create_saved_queries_tab(self):
        """Create a tab for saved queries"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Instructions label
        self.saved_queries_instructions = QLabel(self.get_translation("Saved queries will appear here:"))
        self.saved_queries_instructions.setWordWrap(True)
        layout.addWidget(self.saved_queries_instructions)

        # Create a table to display saved queries
        self.saved_queries_table = QTableWidget()
        self.saved_queries_table.setColumnCount(3)
        self.saved_queries_table.setHorizontalHeaderLabels([
            self.get_translation("Query Name"),
            self.get_translation("Query Text"),
            self.get_translation("Timestamp")
        ])
        self.saved_queries_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.saved_queries_table)

        # Action buttons
        button_layout = QHBoxLayout()

        # Refresh button to reload saved queries
        self.refresh_saved_queries_btn = QPushButton(self.get_translation("Refresh"))
        self.refresh_saved_queries_btn.setObjectName("refresh")  # Blue on hover for refresh
        self.refresh_saved_queries_btn.clicked.connect(self.refresh_saved_queries)
        button_layout.addWidget(self.refresh_saved_queries_btn)

        # Execute selected query button
        self.execute_saved_query_btn = QPushButton(self.get_translation("execute_selected_query"))
        self.execute_saved_query_btn.setObjectName("ok")
        self.execute_saved_query_btn.clicked.connect(self.execute_saved_query)
        button_layout.addWidget(self.execute_saved_query_btn)

        # Delete selected query button
        self.delete_saved_query_btn = QPushButton(self.get_translation("Delete Query"))
        self.delete_saved_query_btn.setObjectName("delete")  # Red for delete
        self.delete_saved_query_btn.clicked.connect(self.delete_saved_query)
        button_layout.addWidget(self.delete_saved_query_btn)

        # Add stretch to push buttons to the left
        button_layout.addStretch()

        layout.addLayout(button_layout)

        # Load saved queries initially
        self.refresh_saved_queries()

        return widget

    def refresh_saved_queries(self):
        """Refresh the saved queries table"""
        if hasattr(self, 'saved_queries_table'):
            # Clear the table
            self.saved_queries_table.setRowCount(0)

            # Get saved queries from the history manager
            saved_queries = self.query_history_manager.get_queries()

            # Populate the table with saved queries
            for i, query_info in enumerate(saved_queries):
                self.saved_queries_table.insertRow(i)

                # For saved queries, the description is the first parameter in add_query
                # and the query is the actual query text
                name_item = QTableWidgetItem(query_info.get('description', ''))
                query_item = QTableWidgetItem(query_info.get('query', ''))
                timestamp_item = QTableWidgetItem(query_info.get('timestamp', ''))

                self.saved_queries_table.setItem(i, 0, name_item)
                self.saved_queries_table.setItem(i, 1, query_item)
                self.saved_queries_table.setItem(i, 2, timestamp_item)

            # Resize columns to fit content
            self.saved_queries_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

    def execute_selected_special_query(self):
        """Execute the selected special query from the combo box - for backward compatibility"""
        if not self.db_manager:
            QMessageBox.warning(self, "Warning", self.get_translation("Please connect to a database first."))
            return

        selected_query_text = self.special_queries_combo.currentText()

        # Dictionary mapping query descriptions to actual SQL queries for the railway station DB
        query_map = {
            "Получить все поезда с их деталями": "SELECT * FROM train",
            "Получить всех сотрудников с их должностями": "SELECT full_name, position FROM employee ORDER BY position",
            "Получить запланированные прибытия и отправления": "SELECT arrival_time, departure_time, train_id, platform_id FROM schedule",
            "Получить пассажиров с билетами": "SELECT p.full_name, p.passport_number, t.ticket_id, t.seat_number FROM passenger p JOIN ticket t ON p.passenger_id = t.passenger_id",
            "Получить назначение сотрудников на поезда": "SELECT e.full_name, e.position, t.train_id FROM employee e JOIN assignment a ON e.employee_id = a.employee_id JOIN train t ON a.train_id = t.train_id",
            "Получить все запланированные поезда с информацией о платформе": "SELECT s.schedule_id, s.arrival_time, s.departure_time, s.date, t.train_id, tr.type, p.platform_id, p.location, p.capacity FROM schedule s JOIN train tr ON s.train_id = tr.train_id JOIN platform p ON s.platform_id = p.platform_id",
            "Получить пассажиров с их билетами и информацией о поезде": "SELECT pass.full_name AS passenger_name, pass.passport_number, t.ticket_id, t.seat_number, tr.train_id, tr.type FROM passenger pass JOIN ticket t ON pass.passenger_id = t.passenger_id JOIN schedule s ON t.ticket_id = s.ticket_id JOIN train tr ON s.train_id = tr.train_id",
            "Получить сотрудников, назначенных на определенные поезда, с их деталями": "SELECT e.employee_id, e.full_name, e.position, e.experience, t.train_id, t.type FROM employee e JOIN assignment a ON e.employee_id = a.employee_id JOIN train t ON a.train_id = t.train_id",
            "Получить услуги с назначениями сотрудникам": "SELECT s.service_name, s.price, s.type, s.service_date, e.full_name AS employee_name, e.position FROM service s JOIN service_assignment sa ON s.service_id = sa.service_id JOIN employee e ON sa.employee_id = e.employee_id",
            "Получить подробное расписание с информацией о пассажирах": "SELECT s.arrival_time, s.departure_time, s.date, tr.train_id, tr.type, p.platform_id, pass.full_name AS passenger_name, t.seat_number FROM schedule s JOIN train tr ON s.train_id = tr.train_id JOIN platform p ON s.platform_id = p.platform_id JOIN ticket t ON s.ticket_id = t.ticket_id JOIN passenger pass ON t.passenger_id = pass.passenger_id",
            "Получить всех пассажиров по поезду": "SELECT tr.type, tr.train_id, pass.full_name, t.seat_number FROM train tr JOIN schedule s ON tr.train_id = s.train_id JOIN ticket t ON s.ticket_id = t.ticket_id JOIN passenger pass ON t.passenger_id = pass.passenger_id ORDER BY tr.train_id",
            "Получить назначения сотрудников по типу поезда": "SELECT tr.type, e.full_name, e.position FROM train tr JOIN assignment a ON tr.train_id = a.train_id JOIN employee e ON a.employee_id = e.employee_id ORDER BY tr.type",
            "Получить использование платформ": "SELECT p.location, p.capacity, COUNT(s.schedule_id) as scheduled_trains FROM platform p LEFT JOIN schedule s ON p.platform_id = s.platform_id GROUP BY p.platform_id, p.location, p.capacity ORDER BY p.platform_id",
            "Получить билеты по диапазону цен": "SELECT t.ticket_id, t.seat_number, t.price, pass.full_name FROM ticket t JOIN passenger pass ON t.passenger_id = pass.passenger_id WHERE t.price BETWEEN 5 AND 10 ORDER BY t.price",
            "Получить запланированные поезда по дате": "SELECT s.date, s.arrival_time, s.departure_time, tr.type, tr.train_id, p.location FROM schedule s JOIN train tr ON s.train_id = tr.train_id JOIN platform p ON s.platform_id = p.platform_id WHERE s.date = '2025-10-27' ORDER BY s.arrival_time",
            "Получить поезда определенного типа": "SELECT * FROM train WHERE type = 'Пассажирский'",
            "Получить сотрудников с определенной должностью": "SELECT * FROM employee WHERE position = 'Машинист'",
            "Получить сотрудников с опытом более 10 лет": "SELECT * FROM employee WHERE experience > 10",
            "Получить пассажиров с определенным номером паспорта": "SELECT * FROM passenger WHERE passport_number = 'АВ1234567'",
            "Получить билеты для определенного номера вагона": "SELECT * FROM ticket WHERE carriage_number = 10",
            "Получить расписание на определенную дату": "SELECT * FROM schedule WHERE date = '2025-10-27'",
            "Получить расписание для определенного идентификатора поезда": "SELECT * FROM schedule WHERE train_id = 1"
        }

        # Parametrized query map - queries that need parameters
        parametrized_query_map = {
            "Получить поезда определенного типа": "SELECT * FROM train WHERE type = ?",
            "Получить сотрудников с определенной должностью": "SELECT * FROM employee WHERE position = ?",
            "Получить сотрудников с опытом более X лет": "SELECT * FROM employee WHERE experience > ?",
            "Получить пассажиров с определенным номером паспорта": "SELECT * FROM passenger WHERE passport_number = ?",
            "Получить билеты для определенного номера вагона": "SELECT * FROM ticket WHERE carriage_number = ?",
            "Получить расписание на определенную дату": "SELECT * FROM schedule WHERE date = ?",
            "Получить расписание для определенного идентификатора поезда": "SELECT * FROM schedule WHERE train_id = ?",
            "Получить услуги определенного типа": "SELECT * FROM service WHERE type = ?",
            "Получить платформы с определенной вместимостью": "SELECT * FROM platform WHERE capacity > ?"
        }

        # Determine if the query is a regular query or a parametrized one
        query = None
        params = None

        if selected_query_text in query_map:
            query = query_map[selected_query_text]
        elif selected_query_text in parametrized_query_map:
            query = parametrized_query_map[selected_query_text]
            params_text = self.query_params_input.text().strip()
            if params_text:
                # Split parameters by comma and strip whitespace
                params = [p.strip() for p in params_text.split(',')]
            else:
                # If this is a parametrized query but no parameters were provided
                QMessageBox.warning(self, "Warning", self.get_translation("parametrized_query_requires_params").format(selected_query=selected_query_text))
                return
        else:
            # Check if it's one of the lab work queries (show a message that the lab queries are organized in groups)
            if "Lab" in selected_query_text or "lab" in selected_query_text:
                QMessageBox.information(self, "Information", "This lab query is now available in the organized lab work query groups. Please use the grouped buttons above.")
                return
            QMessageBox.warning(self, "Warning", self.get_translation("query_not_in_predefined_list"))
            return

        try:
            # Switch to the Query Editor tab to show results
            self.tab_widget.setCurrentIndex(1)  # Index 1 is the query editor tab

            # Display the query in the input field (with parameters if any)
            if params:
                # Format the query showing actual parameter values
                formatted_query = query
                for param in params:
                    formatted_query = formatted_query.replace('?', f"'{param}'", 1)
                self.query_input.setPlainText(formatted_query)
            else:
                self.query_input.setPlainText(query)

            # Execute the query based on whether it needs parameters
            if params:
                result = self.db_manager.execute_query(query, tuple(params))
            else:
                result = self.db_manager.execute_query(query)

            # Add the query to history with a descriptive name
            query_description = f"Special query: {selected_query_text}"
            final_query = query
            if params:
                # Format the query showing actual parameter values
                final_query = query
                for param in params:
                    final_query = final_query.replace('?', f"'{param}'", 1)
            self.query_history_manager.add_query(final_query, query_description)
            self.refresh_query_history()

            if result is not None:  # SELECT query
                if result:
                    # Get column names
                    columns = [desc[0] for desc in self.db_manager.cursor.description]

                    # Set up the results table
                    self.query_results.setRowCount(len(result))
                    self.query_results.setColumnCount(len(columns))
                    self.query_results.setHorizontalHeaderLabels(columns)

                    # Populate the table with color coding for data types
                    for i, row in enumerate(result):
                        for j, value in enumerate(row):
                            item = QTableWidgetItem(str(value))

                            # Determine data type and apply appropriate styling
                            value_str = str(value)
                            if value_str.isdigit() or (value_str.startswith('-') and value_str[1:].isdigit()) or self.is_numeric(value_str):
                                # Numeric data - greenish
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_NUMERIC)
                                item.setBackground(QColor(COLOR_NUMERIC_BG))
                                item.setForeground(QColor(COLOR_NUMERIC_FG))
                            elif self.is_date_or_time(value_str):
                                # Date/time data - orange
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_DATE)
                                item.setBackground(QColor(COLOR_DATE_BG))
                                item.setForeground(QColor(COLOR_DATE_FG))
                            elif value_str.lower() in ['true', 'false', '1', '0', 'yes', 'no']:
                                # Boolean data - teal
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_BOOLEAN)
                                item.setBackground(QColor(COLOR_BOOLEAN_BG))
                                item.setForeground(QColor(COLOR_BOOLEAN_FG))
                            else:
                                # Text data - blueish
                                item.setData(Qt.ItemDataRole.UserRole, DATA_TYPE_TEXT)
                                item.setBackground(QColor(COLOR_TEXT_BG))
                                item.setForeground(QColor(COLOR_TEXT_FG))

                            self.query_results.setItem(i, j, item)

                    # Resize columns to fit content
                    self.query_results.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
                else:
                    # No results but query executed successfully
                    self.query_results.setRowCount(0)
                    self.query_results.setColumnCount(0)

                self.statusBar().showMessage(self.get_translation("special_query_executed_successfully_rows").format(count=len(result) if result else 0))
            else:  # Non-SELECT query
                self.query_results.setRowCount(0)
                self.query_results.setColumnCount(0)
                self.statusBar().showMessage(self.get_translation("special_query_executed_successfully"))

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error executing special query: {str(e)}")

    def execute_saved_query(self):
        """Execute the selected saved query"""
        current_row = self.saved_queries_table.currentRow()
        if current_row >= 0:
            query_item = self.saved_queries_table.item(current_row, 1)  # Query text column
            if query_item:
                query = query_item.text()
                self.execute_query_and_display(query)

    def delete_saved_query(self):
        """Delete the selected saved query"""
        current_row = self.saved_queries_table.currentRow()
        if current_row >= 0:
            query_name_item = self.saved_queries_table.item(current_row, 0)  # Query name column
            if query_name_item:
                query_name = query_name_item.text()

                reply = QMessageBox.question(
                    self,
                    "Confirm Delete",
                    f"Are you sure you want to delete the saved query '{query_name}'?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )

                if reply == QMessageBox.StandardButton.Yes:
                    # Delete the query by its description
                    success = self.query_history_manager.delete_query_by_description(query_name)
                    if success:
                        # The refresh has already been called by the callback in delete_query_by_description
                        pass
                    else:
                        QMessageBox.warning(self, "Warning", f"Query '{query_name}' not found in history.")







    def update_table_columns(self, table_name):
        """Update column selection combo when table changes"""
        if not self.db_manager or not table_name:
            return

        try:
            table_info = self.db_manager.get_table_info(table_name)
            column_names = [col['name'] for col in table_info]



        except Exception as e:
            print(f"Error updating table columns: {e}")
















    def create_relational_model_tab(self):
        """Create a tab to display tables in a relational model format"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Instructions label
        instructions = QLabel("View database tables with their relationships (Relational Model):")
        instructions.setWordWrap(True)
        layout.addWidget(instructions)

        # Create scroll area for the relationship view
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # Create a widget to visualize relationships (will be replaced with a more visual implementation)
        self.relations_view = QLabel("Relationships between tables will be displayed here.\n\n"
                                    "Tables with foreign key relationships will be shown with connecting lines.")
        self.relations_view.setWordWrap(True)
        self.relations_view.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.relations_view.setStyleSheet("QLabel { background-color: #f0f0f0; padding: 20px; border: 1px solid #ccc; }")

        # Create a scrollable widget to show relational model
        self.relations_visualizer = QWidget()
        self.relations_layout = QVBoxLayout(self.relations_visualizer)
        self.relations_visualizer.setStyleSheet("background-color: white;")

        # Create scroll area for the relations visualization
        self.relations_scroll = QScrollArea()
        self.relations_scroll.setWidget(self.relations_visualizer)
        self.relations_scroll.setWidgetResizable(True)
        scroll_layout.addWidget(self.relations_scroll)

        # Hide the old label since we're using the visualizer
        self.relations_view.hide()

        # Add a button to refresh relationships
        refresh_relations_btn = QPushButton("Refresh Relationships")
        refresh_relations_btn.clicked.connect(self.refresh_relationships)
        refresh_relations_btn.setObjectName("refresh")  # Blue on hover for refresh
        scroll_layout.addWidget(refresh_relations_btn)

        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        # Add stretch to push content to the top
        scroll_layout.addStretch()

        return widget

    def refresh_relationships(self):
        """Refresh relationships display"""
        # Clear the existing layout content
        while self.relations_layout.count():
            child = self.relations_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        if self.db_manager:
            try:
                # Get all tables
                tables = self.db_manager.get_tables()
                if not tables:
                    no_tables_label = QLabel("No tables found in the database.")
                    no_tables_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.relations_layout.addWidget(no_tables_label)
                    return

                # Detect relationships between tables
                relationships = self.detect_table_relationships(tables)

                if not relationships:
                    # Also try to infer relationships by naming convention
                    relationships = self.infer_relationships_by_naming(tables)

                # Create visualization of tables and relationships
                self.create_relational_visualization(tables, relationships)

            except Exception as e:
                error_label = QLabel(f"Error detecting relationships: {str(e)}")
                error_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self.relations_layout.addWidget(error_label)
        else:
            no_conn_label = QLabel("No database connection. Please connect to a database first.")
            no_conn_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.relations_layout.addWidget(no_conn_label)

    def create_relational_visualization(self, tables, relationships):
        """Create a visual representation of tables and their relationships"""
        # Create a main layout for the visualization
        main_layout = QVBoxLayout()

        # Add title
        title_label = QLabel("Relational Model View")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; padding: 10px;")
        main_layout.addWidget(title_label)

        # Add description
        desc_label = QLabel("Tables and their relationships in the database:")
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desc_label.setStyleSheet("padding: 5px;")
        main_layout.addWidget(desc_label)

        # Create a horizontal layout to organize tables
        tables_layout = QHBoxLayout()

        # For each table, create a visual representation
        for table in tables:
            table_group = QGroupBox(table)
            table_group.setStyleSheet("QGroupBox { font-weight: bold; border: 1px solid gray; padding: 5px; }")

            table_inner_layout = QVBoxLayout()

            # Get table columns
            try:
                table_info = self.db_manager.get_table_info(table)
                for col in table_info:
                    col_text = f"{col['name']} ({col['type']})"
                    if col['primary_key']:
                        col_text += " [PK]"
                    elif col['name'] in [rel.split('.')[1] for rel in relationships if table in rel.split(' -> ')[0]]:
                        col_text += " [FK]"
                    col_label = QLabel(col_text)
                    table_inner_layout.addWidget(col_label)
            except:
                # If there's an error getting table info, just show the table name
                col_label = QLabel("Error loading columns")
                table_inner_layout.addWidget(col_label)

            table_group.setLayout(table_inner_layout)
            tables_layout.addWidget(table_group)

        # Add the tables layout to main layout
        main_layout.addLayout(tables_layout)

        # Show relationships
        if relationships:
            rel_title = QLabel("Table Relationships:")
            rel_title.setStyleSheet("font-weight: bold; padding-top: 10px;")
            main_layout.addWidget(rel_title)

            rel_list = QListWidget()
            for rel in relationships:
                rel_list.addItem(rel)
            main_layout.addWidget(rel_list)
        else:
            no_rel_label = QLabel("No relationships detected between tables.")
            no_rel_label.setStyleSheet("padding: 10px;")
            main_layout.addWidget(no_rel_label)

        # Add the main layout to the relations layout
        container_widget = QWidget()
        container_widget.setLayout(main_layout)
        self.relations_layout.addWidget(container_widget)

        # Add stretch to use available space
        self.relations_layout.addStretch()

    def detect_table_relationships(self, tables):
        """Detect actual foreign key relationships between tables"""
        relationships = []
        if self.db_manager.db_type == 'postgresql':
            for table in tables:
                # Get foreign key constraints for this table
                try:
                    # Use PostgreSQL-specific query to get foreign key constraints
                    fk_query = """
                        SELECT
                            tc.table_name,
                            kcu.column_name,
                            ccu.table_name AS foreign_table_name,
                            ccu.column_name AS foreign_column_name
                        FROM information_schema.table_constraints AS tc
                        JOIN information_schema.key_column_usage AS kcu
                            ON tc.constraint_name = kcu.constraint_name
                            AND tc.table_schema = kcu.table_schema
                        JOIN information_schema.constraint_column_usage AS ccu
                            ON ccu.constraint_name = tc.constraint_name
                            AND ccu.table_schema = tc.table_schema
                        WHERE tc.constraint_type = 'FOREIGN KEY' AND tc.table_name = %s;
                    """
                    self.db_manager.cursor.execute(fk_query, (table,))
                    fk_constraints = self.db_manager.cursor.fetchall()

                    for fk in fk_constraints:
                        # fk is a tuple: (table_name, column_name, foreign_table_name, foreign_column_name)
                        if len(fk) >= 4:
                            fk_from_col = fk[1]  # The column in the current table that references another
                            fk_to_table = fk[2]  # The referenced table
                            fk_to_col = fk[3]  # The referenced column

                            if fk_to_table in tables:  # Only if referenced table exists in the database
                                relationships.append(f"{table}.{fk_from_col} -> {fk_to_table}.{fk_to_col}")

        return relationships

    def infer_relationships_by_naming(self, tables):
        """Infer relationships based on naming conventions (e.g., user_id references user table)"""
        relationships = []

        for table in tables:
            table_info = self.db_manager.get_table_info(table)

            for column in table_info:
                col_name = column['name']

                # Check if column name follows convention like tablename_id
                # which likely references another table
                if col_name.endswith('_id'):
                    potential_ref_table = col_name[:-3]  # Remove "_id" suffix

                    # Check if a table with that name exists
                    if potential_ref_table in tables and potential_ref_table != table:
                        # Check if the potential reference table has an 'id' primary key
                        ref_table_info = self.db_manager.get_table_info(potential_ref_table)
                        has_id_pk = any(col['name'] == 'id' and col['primary_key'] for col in ref_table_info)

                        if has_id_pk:
                            relationships.append(f"{table}.{col_name} -> {potential_ref_table}.id")

        return relationships

    def create_special_queries_tab(self):
        """Create a tab for special predefined queries from lab work"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Instructions label
        self.special_queries_instructions = QLabel(self.get_translation("Select and execute predefined queries from lab work #5 and #6:"))
        self.special_queries_instructions.setWordWrap(True)
        layout.addWidget(self.special_queries_instructions)

        # Create scroll area for query selection
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # Create group for lab work 5 queries
        lab5_group = QGroupBox("Laboratory Work 5 - SELECT Queries")
        lab5_layout = QVBoxLayout()

        # Lab 5 queries buttons
        lab5_queries = [
            ("1. Trains with speed > 100", LabQueries.LAB5_QUERY1),
            ("2. Train types sorted by speed", LabQueries.LAB5_QUERY2),
            ("3. Trains with 'Пассажирский' or speed > 140", LabQueries.LAB5_QUERY3),
            ("4. Carriage number and price of all tickets", LabQueries.LAB5_QUERY4),
            ("5. Tickets with price < 6", LabQueries.LAB5_QUERY5),
            ("6. Tickets sorted by descending price", LabQueries.LAB5_QUERY6),
            ("7. Tickets with price 3 < x <= 7", LabQueries.LAB5_QUERY7),
            ("8. Service name and total price", LabQueries.LAB5_QUERY8),
            ("9. Materials with price > 1000 and date before 2025-01-20", LabQueries.LAB5_QUERY9),
            ("10. Services sorted by ascending date", LabQueries.LAB5_QUERY10),
            ("11. Services with price > 1000 and type 'Техническая'", LabQueries.LAB5_QUERY11),
            ("12. Carriage enumeration and departure time", LabQueries.LAB5_QUERY12),
            ("13. Carriage enumerations after 12:00:00", LabQueries.LAB5_QUERY13),
            ("14. Carriage enumerations sorted by departure time desc", LabQueries.LAB5_QUERY14),
            ("15. Difference between arrival and departure > 20 minutes", LabQueries.LAB5_QUERY15),
            ("16. First 5 records about platforms", LabQueries.LAB5_QUERY16),
            ("17. Platforms with capacity 300-500", LabQueries.LAB5_QUERY17),
            ("18. Tickets sorted by location then capacity", LabQueries.LAB5_QUERY18),
            ("19. Platforms with capacity > 400 or track > 3", LabQueries.LAB5_QUERY19),
            ("20. First 10 passengers: passport, name, phone", LabQueries.LAB5_QUERY20),
            ("21. Passengers with ID 5 < x < 10", LabQueries.LAB5_QUERY21),
            ("22. Passengers sorted by full name reverse", LabQueries.LAB5_QUERY22),
            ("23. Passengers with first letter 'A' or phone +37544", LabQueries.LAB5_QUERY23),
            ("24. First 5 employees with experience > 10", LabQueries.LAB5_QUERY24),
            ("25. Only engineers", LabQueries.LAB5_QUERY25),
            ("26. Employees sorted by passport desc", LabQueries.LAB5_QUERY26),
            ("27. All positions except 'Проводник'", LabQueries.LAB5_QUERY27),
            ("28. Tickets where carriage number = train carriage count", LabQueries.LAB5_QUERY28),
            ("29. All trains with tickets for last carriage", LabQueries.LAB5_QUERY29),
            ("30. All tickets > 5 with last carriages", LabQueries.LAB5_QUERY30),
            ("31. Non-freight trains with tickets > 5 and carriage > 12", LabQueries.LAB5_QUERY31),
            ("32. Tickets > 5, non-freight, for one carriage", LabQueries.LAB5_QUERY32),
            ("33. Services with employees who provide them", LabQueries.LAB5_QUERY33),
            ("34. Employees with positions and services", LabQueries.LAB5_QUERY34),
            ("35. All services and employees providing them", LabQueries.LAB5_QUERY35),
            ("36. Full join of services and employees", LabQueries.LAB5_QUERY36),
            ("37. Cartesian product of all services and employees", LabQueries.LAB5_QUERY37),
            ("38. Platform location with times and track > 2", LabQueries.LAB5_QUERY38),
            ("39. Schedules with track locations capacity > 400", LabQueries.LAB5_QUERY39),
            ("40. Departure time of train 2 with platform", LabQueries.LAB5_QUERY40),
            ("41. All data about platform 4", LabQueries.LAB5_QUERY41),
            ("42. Cartesian product Platform 1 Minsk with schedules", LabQueries.LAB5_QUERY42),
            ("43. Services provided by drivers and names", LabQueries.LAB5_QUERY43),
            ("44. All drivers and services they provide", LabQueries.LAB5_QUERY44),
            ("45. All technical services and employees", LabQueries.LAB5_QUERY45),
            ("46. All engineers providing technical services", LabQueries.LAB5_QUERY46),
            ("47. Cartesian product services and specific employee", LabQueries.LAB5_QUERY47),
            ("48. Trains speed > 100 with arrival/departure times", LabQueries.LAB5_QUERY48),
            ("49. Arrival/departure for trains speed > 100", LabQueries.LAB5_QUERY49),
            ("50. Arrival/departure for trains speed > 100 (with/without schedule)", LabQueries.LAB5_QUERY50),
            ("51. Full join express trains and departure < 12:00", LabQueries.LAB5_QUERY51),
            ("52. Cartesian product express trains with schedules", LabQueries.LAB5_QUERY52),
            ("53. Seat numbers and passenger names with price > 7", LabQueries.LAB5_QUERY53),
            ("54. Ticket cost > 7 with passengers on seats", LabQueries.LAB5_QUERY54),
            ("55. Seat and names of passengers tickets > 7", LabQueries.LAB5_QUERY55),
            ("56. Full join tickets and passengers with discount price > 7", LabQueries.LAB5_QUERY56),
            ("57. Cartesian product all tickets and passengers > 7", LabQueries.LAB5_QUERY57),
            ("58. Ticket data and departure time > 7", LabQueries.LAB5_QUERY58),
            ("59. All schedule data and ticket data > 7", LabQueries.LAB5_QUERY59),
            ("60. All tickets with seats and times before 12:00", LabQueries.LAB5_QUERY60),
            ("61. Full join tickets < 7 and times after 12:00", LabQueries.LAB5_QUERY61),
            ("62. Cartesian product tickets > 7 and times before 12:00", LabQueries.LAB5_QUERY62)
        ]

        # Create buttons for Lab 5 queries
        for query_desc, query in lab5_queries:
            btn = QPushButton(query_desc)
            btn.clicked.connect(lambda checked, q=query: self.execute_specific_query(q))
            lab5_layout.addWidget(btn)

        lab5_group.setLayout(lab5_layout)
        scroll_layout.addWidget(lab5_group)

        # Create group for lab work 6 queries
        lab6_group = QGroupBox("Laboratory Work 6 - Aggregate, Subquery, Set Operations")
        lab6_layout = QVBoxLayout()

        # Lab 6 queries buttons
        lab6_queries = [
            ("1. Average speed and count of year of manufacture", LabQueries.LAB6_QUERY1),
            ("2. Total sum of ticket costs, min seat, max carriage", LabQueries.LAB6_QUERY2),
            ("3. Unique services with avg cost > 1000 and earliest date", LabQueries.LAB6_QUERY3),
            ("4. Values matching patterns with 10th hour and dates", LabQueries.LAB6_QUERY4),
            ("5. Union of Platform and Train tables", LabQueries.LAB6_QUERY5),
            ("6. Intersect of Passenger table selections", LabQueries.LAB6_QUERY6),
            ("7. Except: unique rows from first excluding second", LabQueries.LAB6_QUERY7)
        ]

        # Create buttons for Lab 6 queries
        for query_desc, query in lab6_queries:
            btn = QPushButton(query_desc)
            btn.clicked.connect(lambda checked, q=query: self.execute_specific_query(q))
            lab6_layout.addWidget(btn)

        lab6_group.setLayout(lab6_layout)
        scroll_layout.addWidget(lab6_group)

        # Add empty combo box for backward compatibility (not visible to user but exists for the function)
        self.special_queries_combo = QComboBox()
        self.special_queries_combo.hide()  # Hide the combo box since we're using organized buttons
        scroll_layout.addWidget(self.special_queries_combo)

        # Parameters input (for potential backward compatibility)
        self.query_params_input = QLineEdit()
        self.query_params_input.setToolTip("Enter parameters for parametrized queries (comma-separated)")
        self.query_params_input.hide()  # Hide since our queries are not parameterized
        scroll_layout.addWidget(self.query_params_input)

        # Execute button (for backward compatibility)
        self.execute_special_query_btn = QPushButton(self.get_translation("execute_selected_query"))
        self.execute_special_query_btn.setObjectName("ok")  # Green on hover for execute
        self.execute_special_query_btn.clicked.connect(self.execute_selected_special_query)
        self.execute_special_query_btn.hide()  # Hide since we have direct buttons for each query
        scroll_layout.addWidget(self.execute_special_query_btn)

        # Save results button
        self.save_special_results_btn = QPushButton(self.get_translation("save_results_excel"))
        self.save_special_results_btn.setObjectName("save")  # Green on hover for save
        self.save_special_results_btn.clicked.connect(self.export_query_results)
        scroll_layout.addWidget(self.save_special_results_btn)

        # Add some stretch to push content to top
        scroll_layout.addStretch()

        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        return widget


    def apply_data_type_highlighting(self, item, value):
        """Apply background and text colors based on data type"""
        from utils.constants import COLOR_NUMERIC_BG, COLOR_NUMERIC_FG, COLOR_DATE_BG, COLOR_DATE_FG, COLOR_BOOLEAN_BG, COLOR_BOOLEAN_FG, COLOR_TEXT_BG, COLOR_TEXT_FG

        # Check if value is numeric
        if self.is_numeric(value):
            item.setBackground(QColor(COLOR_NUMERIC_BG))
            item.setForeground(QColor(COLOR_NUMERIC_FG))
        # Check if value is date/time
        elif self.is_date_or_time(value):
            item.setBackground(QColor(COLOR_DATE_BG))
            item.setForeground(QColor(COLOR_DATE_FG))
        # Check if value is boolean
        elif value.lower() in ['true', 'false', '1', '0']:
            item.setBackground(QColor(COLOR_BOOLEAN_BG))
            item.setForeground(QColor(COLOR_BOOLEAN_FG))
        else:
            # Default text color
            item.setBackground(QColor(COLOR_TEXT_BG))
            item.setForeground(QColor(COLOR_TEXT_FG))

    def clear_layout(self, layout):
        """Recursively clear all widgets from a layout"""
        while layout.count():
            child = layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
            elif child.layout():
                self.clear_layout(child.layout())